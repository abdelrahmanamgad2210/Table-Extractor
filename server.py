"""
Table Extractor UI  —  stateless cloud-ready architecture
──────────────────────────────────────────────────────────
The client holds the accumulated Excel blob in memory.
Each /upload receives: new document + current Excel (if any).
Server returns: updated Excel as binary + row counts in headers.
This means ZERO server-side state — works on any cloud platform.

Local:  python server.py  →  http://localhost:5000
Cloud:  set GEMINI_API_KEY env var, deploy.
"""
import base64, json, sys, os, re, time, urllib.request, urllib.error, logging, hashlib
from io import BytesIO
from pathlib import Path
from datetime import datetime, timezone

# ── Config ────────────────────────────────────────────────────────────────────
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    try:
        _text = _env_path.read_text(encoding="utf-8-sig")  # utf-8-sig strips BOM
    except UnicodeDecodeError:
        _text = _env_path.read_text(encoding="utf-16")
    for _line in _text.splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _, _v = _line.partition("=")
            _v = _v.strip().strip('"').strip("'")  # remove surrounding quotes
            os.environ.setdefault(_k.strip(), _v)

API_KEY  = os.environ.get("GEMINI_API_KEY", "")
MODEL    = os.environ.get("GEMINI_MODEL",   "gemini-2.5-flash")

if not API_KEY:
    print("\n  ERROR: GEMINI_API_KEY is not set.")
    print("  Add this line to .env (no quotes):")
    print("    GEMINI_API_KEY=AIzaSy...")
    print("  Get a free key at: https://aistudio.google.com/apikey\n")
    sys.exit(1)

print(f"  API key loaded: {API_KEY[:8]}...{API_KEY[-4:]}")
IS_LOCAL = not os.environ.get("PORT")

# ── Logging ───────────────────────────────────────────────────────────────────
_log_file = Path(__file__).parent / "server.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(_log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)
PORT     = int(os.environ.get("PORT", 5000))

# Local backup path (only used when running locally)
LOCAL_EXCEL = Path(r"C:\Users\Abood\Downloads\Alkhaleej_extracted.xlsx")

# ── Cache ─────────────────────────────────────────────────────────────────────
CACHE_DIR = Path(__file__).parent / "cache"
CACHE_DIR.mkdir(exist_ok=True)

def _cache_get(key: str):
    p = CACHE_DIR / f"{key}.json"
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else None

def _cache_set(key: str, val):
    (CACHE_DIR / f"{key}.json").write_text(json.dumps(val, ensure_ascii=False), encoding="utf-8")

# ── Dashboard Row Store ───────────────────────────────────────────────────────
DATA_DIR   = Path(__file__).parent / "data"
STORE_PATH = DATA_DIR / "row_store.json"
DATA_DIR.mkdir(exist_ok=True)

def load_row_store() -> dict:
    if not STORE_PATH.exists():
        return {"created_at": datetime.now(timezone.utc).isoformat(), "rows": []}
    with STORE_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_row_store(store: dict):
    store["last_modified"] = datetime.now(timezone.utc).isoformat()
    with STORE_PATH.open("w", encoding="utf-8") as f:
        json.dump(store, f, ensure_ascii=False, indent=2)

# ── Correction Memory (Section 10) ────────────────────────────────────────────
CORRECTION_PATH = DATA_DIR / "correction_memory.json"

def load_corrections() -> dict:
    if not CORRECTION_PATH.exists():
        return {"corrections": [], "confirmed_values": {}}
    with CORRECTION_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_corrections(mem: dict):
    DATA_DIR.mkdir(exist_ok=True)
    with CORRECTION_PATH.open("w", encoding="utf-8") as f:
        json.dump(mem, f, ensure_ascii=False, indent=2)

def _char_diffs(old: str, new: str) -> list:
    if not old or not new or len(old) != len(new):
        return []
    return [{"pos": i, "from": a, "to": b}
            for i, (a, b) in enumerate(zip(old, new)) if a != b]

def record_correction(source_file: str, field: str,
                      model_read: str, human_value: str, reason: str):
    """Record a human correction so it can be injected as a few-shot example."""
    if model_read == human_value:
        return
    authority = detect_authority(source_file)
    mem = load_corrections()
    next_id = max((c["_id"] for c in mem["corrections"]), default=0) + 1
    mem["corrections"].append({
        "_id":               next_id,
        "source_file":       source_file,
        "issuing_authority": authority,
        "field":             field,
        "model_read":        model_read,
        "human_value":       human_value,
        "diff_chars":        _char_diffs(model_read, human_value),
        "reason_flagged":    reason,
        "corrected_at":      datetime.now(timezone.utc).isoformat(),
    })
    cv = mem["confirmed_values"].setdefault(field, [])
    if human_value and human_value not in cv:
        cv.append(human_value)
    save_corrections(mem)
    log.info("Correction recorded: %s '%s' → '%s'", field, model_read, human_value)

def apply_confirmed_value_boost(row: dict, flags: dict) -> dict:
    """
    If a freshly-read critical field exactly equals a previously human-confirmed
    value, drop the flag for that field. NEVER fills empty reads.
    """
    confirmed = load_corrections().get("confirmed_values", {})
    for field in ("الملاحظات", "اللوحة"):
        val = str(row.get(field, "") or "").strip()
        if val and val in confirmed.get(field, []) and field in flags:
            log.info("Confirmed-value boost: removed flag on %s = '%s'", field, val)
            flags.pop(field)
    return flags

def build_fewshot_block(authority: str = "unknown") -> str:
    """Render the most relevant past corrections as a prompt-injectable block."""
    from collections import Counter
    ranked = select_fewshot_examples(authority)
    if not ranked:
        return ""
    lines = [
        "## LEARNED CORRECTIONS FROM PREVIOUS HUMAN REVIEWS",
        "Human reviewers corrected these misreads on similar documents. "
        "Use them to read MORE CAREFULLY in the same spots — but still transcribe "
        "only what you actually see. Do NOT copy these values blindly.\n",
    ]
    pair_counts: Counter = Counter()
    for c in ranked:
        diffs = c.get("diff_chars", [])
        if diffs:
            spots = ", ".join(f"pos {d['pos']+1}: '{d['from']}'→'{d['to']}'" for d in diffs)
            lines.append(
                f"- {c['field']}: misread \"{c['model_read']}\" → correct \"{c['human_value']}\" "
                f"({spots}). Reason: {c['reason_flagged']}."
            )
            for d in diffs:
                pair_counts[f"{d['from']}↔{d['to']}"] += 1
        else:
            lines.append(
                f"- {c['field']}: misread \"{c['model_read']}\" → correct \"{c['human_value']}\" "
                f"({c['reason_flagged']})."
            )
    if pair_counts:
        top = ", ".join(f"{p} (×{n})" for p, n in pair_counts.most_common(5))
        lines.append(f"\nMost frequent confusion pairs: {top}. Double-check these characters.")
    return "\n".join(lines)

# ── Source Profiles & Authority Detection (Section 10.3) ─────────────────────
PROFILE_PATH = DATA_DIR / "source_profiles.json"

_DEFAULT_PROFILES = {
    "dubai_police": {
        "layout": "B", "column_groups": 4, "reading_order": "rtl",
        "vin_column_present": True,
        "common_empty_markers": ["بدون لوحة", "بدون رقم"],
        "date_format": "YYYY-MM-DD",
        "typical_vin_confidence": "medium",
        "notes": "Tiny ~5pt text. Over-flag VINs. Process one column-group at a time."
    },
    "ajman_security": {
        "layout": "A", "column_groups": 1, "reading_order": "rtl",
        "vin_in_notes_column": True,
        "date_format": "D/M/YYYY",
        "typical_vin_confidence": "high",
        "notes": "Clean single table. VIN lives in الملاحظات column."
    },
    "sharjah_police": {
        "layout": "C", "column_groups": 3, "reading_order": "rtl",
        "date_format": "D/M/YYYY",
        "typical_vin_confidence": "low",
        "notes": "Heavy compression, 400+ rows. Many blank VINs. Default confidence to low."
    },
    "ajman_municipality": {
        "layout": "D", "column_groups": 2, "reading_order": "rtl",
        "common_empty_markers": ["***", "بدون لوحة"],
        "date_format": "D/M/YYYY",
        "typical_vin_confidence": "medium",
        "notes": "Watch for *** masked VINs. Reproduce mask exactly."
    },
    "emirates_auction": {
        "layout": "B", "column_groups": 3, "reading_order": "rtl",
        "date_format": "YYYY-MM-DD",
        "typical_vin_confidence": "medium",
        "notes": "Multi-group landscape, similar to Dubai Police."
    },
}

_AUTHORITY_KEYWORDS = {
    "dubai_police":       ["دبي", "dubai", "dxb", "شرطة دبي"],
    "sharjah_police":     ["الشارقة", "sharjah", "shj", "شرطة الشارقة"],
    "ajman_security":     ["عجمان", "ajman", "خدمات أمنية", "أمن عجمان"],
    "ajman_municipality": ["بلدية عجمان", "ajman municipality"],
    "emirates_auction":   ["الإمارات للمزادات", "emirates auction", "auction"],
}

def load_profiles() -> dict:
    if PROFILE_PATH.exists():
        with PROFILE_PATH.open("r", encoding="utf-8") as f:
            return json.load(f)
    profiles = _DEFAULT_PROFILES.copy()
    DATA_DIR.mkdir(exist_ok=True)
    with PROFILE_PATH.open("w", encoding="utf-8") as f:
        json.dump(profiles, f, ensure_ascii=False, indent=2)
    return profiles

def detect_authority(filename: str) -> str:
    """Best-effort authority detection from filename."""
    name = filename.lower().replace("_", " ").replace("-", " ")
    for authority, keywords in _AUTHORITY_KEYWORDS.items():
        for kw in keywords:
            if kw in name:
                return authority
    return "unknown"

def build_source_profile_block(authority: str) -> str:
    profiles = load_profiles()
    p = profiles.get(authority)
    if not p:
        return ""
    markers = ", ".join(p.get("common_empty_markers", []))
    return (
        f"## KNOWN PROFILE FOR THIS SOURCE ({authority.replace('_', ' ').title()})\n"
        f"- Layout: {p.get('layout')}, {p.get('column_groups')} column-group(s), "
        f"{p.get('reading_order','rtl').upper()} reading order\n"
        f"- Date format in source: {p.get('date_format')}\n"
        f"- Default VIN confidence ceiling: {p.get('typical_vin_confidence')}\n"
        + (f"- Empty markers to preserve literally: {markers}\n" if markers else "")
        + f"- Notes: {p.get('notes', '')}"
    )

def select_fewshot_examples(authority: str, max_examples: int = 6) -> list:
    mem  = load_corrections()
    corr = mem.get("corrections", [])
    if not corr:
        return []
    same  = [c for c in corr if c.get("issuing_authority") == authority]
    rest  = [c for c in corr if c.get("issuing_authority") != authority]
    def _rank(c):
        return len(c.get("diff_chars", [])) > 0
    return (sorted(same, key=_rank, reverse=True) +
            sorted(rest, key=_rank, reverse=True))[:max_examples]

def build_extraction_prompt(authority: str) -> str:
    """Assemble base prompt + source profile + few-shot corrections."""
    profile = build_source_profile_block(authority)
    fewshot = build_fewshot_block(authority)
    parts   = [PROMPT]
    if profile:
        parts.append(profile)
    if fewshot:
        parts.append(fewshot)
    return "\n\n".join(parts)

# ── Validation & Flagging Engine ──────────────────────────────────────────────
_VIN_PATTERN   = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$")
_PLATE_PATTERN = re.compile(r"^\d{2,7}$")   # UAE plates: 2–7 digits; 8+ digits → likely رقم التخزين
_PLATE_NO_PLATE = {"بدون لوحة", "بدون رقم", "بدون", "no plate"}  # valid Arabic "no plate" strings
_DATE_PATTERNS = [
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$"),
    re.compile(r"^\d{1,2}-\d{1,2}-\d{4}$"),
    re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$"),
]
_VALIDATED_FIELDS = {
    "الملاحظات":   "vin",
    "اللوحة":      "plate",
    "تاريخ الحجز": "date",
}

def _vld_vin(v: str):
    v = v.strip().upper().replace(" ", "").replace("-", "")
    if not v: return False, "VIN is empty"
    if len(v) != 17: return False, f"VIN length {len(v)} ≠ 17"
    if not _VIN_PATTERN.match(v): return False, "VIN contains invalid chars (I/O/Q not allowed)"
    return True, ""

def _vld_plate(v: str):
    v = v.strip()
    if not v: return False, "Plate is empty"
    if v.lower() in _PLATE_NO_PLATE: return True, ""   # valid "no plate" designations
    if v.isdigit() and len(v) >= 8:
        return False, f"Plate has {len(v)} digits — likely رقم التخزين, not a plate number"
    if not _PLATE_PATTERN.match(v): return False, f'Plate "{v}" is not a valid UAE plate'
    return True, ""

def _vld_date(v: str):
    v = v.strip()
    if not v: return False, "Date is empty"
    for p in _DATE_PATTERNS:
        if p.match(v): return True, ""
    return False, f'Date "{v}" is unrecognised format'

def run_validation_flags(row: dict) -> dict:
    """Return {field: reason} for every validated field that fails its check."""
    flags = {}
    for field, val in row.items():
        if field.startswith("_"):
            continue
        vtype = _VALIDATED_FIELDS.get(field)
        if not vtype:
            continue
        v = str(val or "").strip()
        if vtype == "vin":
            ok, reason = _vld_vin(v)
        elif vtype == "plate":
            ok, reason = _vld_plate(v)
        else:
            ok, reason = _vld_date(v)
        if not ok:
            flags[field] = reason
    return flags

def _detect_table_boundary(rows: list, headers: list) -> int:
    """
    Return the index of the first row that belongs to a SECOND table inside this
    tile.  A boundary is detected when the م serial number resets: after seeing
    a run of serials that reach at least 5, if a later value drops below half
    the highest seen so far we treat it as a new-table reset and stop there.

    Returns len(rows) if no boundary is found (common case).
    """
    # find م column index
    serial_idx = next((i for i, h in enumerate(headers) if h == "م"), None)
    if serial_idx is None:
        return len(rows)

    highest = 0
    for i, row in enumerate(rows):
        raw = str(row[serial_idx] if serial_idx < len(row) else "").strip()
        if not raw.isdigit():
            continue
        n = int(raw)
        if highest >= 5 and n < highest // 2:
            log.warning(
                "Table boundary detected at row %d: م reset from %d → %d. "
                "Truncating %d hallucinated tail rows.",
                i, highest, n, len(rows) - i,
            )
            return i
        if n > highest:
            highest = n
    return len(rows)


def merge_extracted_rows(table: dict, source_filename: str) -> int:
    """Append extracted rows to the persistent row store, merging Gemini flags."""
    store   = load_row_store()
    next_id = max((r["_id"] for r in store["rows"]), default=0) + 1
    headers      = table.get("headers", [])
    gemini_flags = table.get("_gemini_flags", [])   # per-row flags from Gemini

    raw_rows = table.get("rows", [])
    boundary = _detect_table_boundary(raw_rows, headers)
    if boundary < len(raw_rows):
        raw_rows     = raw_rows[:boundary]
        gemini_flags = gemini_flags[:boundary]

    new_rows = []
    for ri, raw_row in enumerate(raw_rows):
        row_dict = {h: str(v or "").strip() for h, v in zip(headers, raw_row)}
        # Merge validation flags: start from Gemini's _needs_review then overlay rule-based checks
        g_flags = gemini_flags[ri] if ri < len(gemini_flags) else {}
        val_flags = run_validation_flags(row_dict)
        # g_flags is {field: reason_string} — use directly; val_flags wins on overlap
        merged_flags = {**g_flags, **val_flags}
        merged_flags = apply_confirmed_value_boost(row_dict, merged_flags)

        row_dict["_id"]           = next_id
        row_dict["source_file"]   = source_filename
        row_dict["_extracted_at"] = datetime.now(timezone.utc).isoformat()
        row_dict["_flags"]        = merged_flags
        row_dict["_all_reviewed"] = False
        new_rows.append(row_dict)
        next_id += 1

    store["rows"].extend(new_rows)
    save_row_store(store)
    log.info("Row store: +%d rows (total %d) from %s", len(new_rows), len(store["rows"]), source_filename)
    return len(new_rows)

# ── Session stats (resets on server restart) ──────────────────────────────────
_stats = {"files": 0, "api_calls": 0, "cache_hits": 0, "frames_skipped": 0, "rows": 0}

# ── Rate limiter (stay safely under 15 RPM free tier) ────────────────────────
_last_call_time = 0.0
API_DELAY = 4.1

def _rate_limit():
    global _last_call_time
    wait = API_DELAY - (time.time() - _last_call_time)
    if wait > 0:
        log.info("Rate limit: waiting %.1fs", wait)
        time.sleep(wait)
    _last_call_time = time.time()

# ── Image utilities ───────────────────────────────────────────────────────────
from PIL import Image, ImageFile as _PIL_ImageFile
_PIL_ImageFile.LOAD_TRUNCATED_IMAGES = True   # survive broken/large GIF data streams

MAX_PX            = 1400   # longest side budget per tile — enough for Arabic numerals, saves ~40% input tokens vs 2000
JPEG_Q            = 75     # JPEG quality — 75 is visually lossless for scanned docs, cuts image byte size vs 90
TILE_PX           = 700    # tile height in original pixels before any scaling
COL_SPLIT_PX      = 1300   # target column-group width — documents wider than this are split
HEADER_PX         = 80     # header strip height to prepend to non-first tiles within a segment
TABLE_HEADER_MIN_PX = 20   # a dark band ≥ this many consecutive rows = a new-table header
TABLE_HEADER_DARK   = 120  # average pixel brightness below this = "dark row"
                            # covers dark-blue (#1F4E79 avg≈77), black (0), and dark-grey headers
                            # but NOT medium-grey cell borders (typically 2–11px and avg>120)

def _enhance(img):
    """Sharpen and boost contrast — critical for dense Arabic numerals."""
    from PIL import ImageFilter, ImageEnhance
    img = ImageEnhance.Contrast(img).enhance(1.4)
    img = ImageEnhance.Sharpness(img).enhance(2.5)
    return img.filter(ImageFilter.SHARPEN)

def _to_jpeg(img) -> bytes:
    buf = BytesIO()
    # subsampling=2 (4:2:0) is the universal-compatible chroma setting and
    # avoids the fileno() failure that "optimize=True + subsampling=0" triggers
    # in Pillow 11+ on Python 3.14 when writing to BytesIO.
    img.save(buf, format="JPEG", quality=JPEG_Q, subsampling=2)
    return buf.getvalue()

def _scale_tile(img):
    """
    Scale a single tile so its longest side fits within MAX_PX.
    When a tile is already under budget (typical for column-strip tiles at
    original resolution) this is a no-op — the native resolution is preserved.
    """
    from PIL import Image
    w, h = img.size
    if max(w, h) <= MAX_PX:
        return img
    scale = MAX_PX / max(w, h)
    return img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

_MIN_DATA_BETWEEN_TABLES = 200   # minimum px of data rows between two table headers

def _scan_dark_bands(img):
    """
    Internal helper: returns a list of (band_start, band_end, band_height) tuples
    for every consecutive run of dark rows in the image.
    Uses row-average grayscale brightness < TABLE_HEADER_DARK for at least
    TABLE_HEADER_MIN_PX consecutive rows.
    """
    w, h = img.size
    sample_w = min(w, 300)
    thin     = img.resize((sample_w, h), Image.BOX).convert("L")
    tw       = sample_w
    px       = thin.load()
    row_avg  = [sum(px[x, y] for x in range(tw)) / tw for y in range(h)]

    dark = [y for y, b in enumerate(row_avg) if b < TABLE_HEADER_DARK]
    bands = []
    if not dark:
        return bands

    s = dark[0]; p = dark[0]
    for y in dark[1:]:
        if y - p > 4:
            # s >= 30: ignore dark bands touching the very top (page border / trim marks)
            if p - s + 1 >= TABLE_HEADER_MIN_PX and s >= 30:
                bands.append((s, p, p - s + 1))
            s = y
        p = y
    if p - s + 1 >= TABLE_HEADER_MIN_PX and s >= 30:
        bands.append((s, p, p - s + 1))
    return bands


def _find_table_layout(img) -> tuple:
    """
    Analyse a column strip and return (first_header_y, boundaries).

    first_header_y: y of the FIRST table's column-header row (0 if the strip
                    starts with the header at the very top, or the first thick
                    dark band if there is a title section above it).

    boundaries:     list of y positions where ADDITIONAL tables start.
                    The first dark band is the first table's header — NOT a
                    boundary.  Only bands appearing at least _MIN_DATA_BETWEEN_TABLES
                    pixels after the previous table's header are returned.

    This ensures:
    • Pre-table title content and the first table share one segment.
    • The correct column header (not y=0 title text) is prepended to non-first
      tiles within the first table's segment.
    • Each subsequent table gets its own segment starting at its header row.
    """
    bands = _scan_dark_bands(img)
    if not bands:
        return 0, []

    first_header_y = bands[0][0]
    boundaries     = []
    prev_end       = bands[0][1]

    for start, end, bh in bands[1:]:
        if start >= prev_end + _MIN_DATA_BETWEEN_TABLES:
            boundaries.append(start)
            prev_end = end

    if boundaries:
        log.info("Table layout: first header y=%d, boundaries=%s", first_header_y, boundaries)
    return first_header_y, boundaries


def _horizontal_tiles(img, forced_cuts=None, first_header_y=0) -> list:
    """
    Split a column strip into overlapping horizontal tiles, respecting table boundaries.

    forced_cuts:     y-coordinates where a NEW table starts (from _find_table_layout).
    first_header_y:  y of the first table's column-header row within the strip.
                     Non-first tiles in the first segment prepend rows starting
                     at this y (the actual column labels) rather than y=0
                     (which might be blank title text above the table).

    Within each segment:
      • Segment 0 (first table): header to prepend = img[first_header_y : first_header_y+HEADER_PX]
      • Segment 1+ (subsequent tables): header to prepend = img[seg_start : seg_start+HEADER_PX]
        (the boundary y IS the start of that table's dark header row)
    """
    w, h   = img.size
    overlap = int(TILE_PX * 0.12)

    cuts     = sorted({0} | {c for c in (forced_cuts or []) if 0 < c < h})
    segments = list(zip(cuts, cuts[1:] + [h]))

    tiles = []
    for si, (seg_start, seg_end) in enumerate(segments):
        if seg_end <= seg_start:
            continue

        # Choose which rows to use as the column-header strip for this segment
        if si == 0:
            hdr_y = max(first_header_y, seg_start)   # actual header, not title-page top
        else:
            hdr_y = seg_start                          # boundary row IS this table's header

        hdr_h    = min(HEADER_PX, seg_end - hdr_y)
        hdr_crop = img.crop((0, hdr_y, w, hdr_y + hdr_h))

        y = seg_start
        while y < seg_end:
            bottom    = min(y + TILE_PX, seg_end)
            data_crop = img.crop((0, y, w, bottom))

            if y == seg_start:
                tile = data_crop              # first tile of segment: header naturally present
            else:
                combined = Image.new("RGB", (w, hdr_h + (bottom - y)), (255, 255, 255))
                combined.paste(hdr_crop,  (0, 0))
                combined.paste(data_crop, (0, hdr_h))
                tile = combined

            tiles.append(tile)
            if bottom == seg_end:
                break
            y = bottom - overlap

    return tiles


def _smart_tiles(img) -> list:
    """
    Main tiling strategy for large multi-column Arabic documents.

    Step 1 — Vertical column-group split:
        Wide documents (Dubai Police, Sharjah, etc.) have 2-4 column groups
        arranged side-by-side RTL.  Splitting each group into its own strip
        gives that group the FULL MAX_PX resolution budget.

    Step 2 — Table-layout analysis:
        Each column strip is scanned for thick dark header bands.  The first
        band locates the first table's column headers; subsequent bands that
        follow at least _MIN_DATA_BETWEEN_TABLES px of data are treated as the
        start of additional tables.  Tile cuts are forced at those boundaries
        so two tables NEVER share a tile.

    Step 3 — Horizontal tiling per segment:
        Each table segment is tiled into TILE_PX-high strips with 12% overlap.
        Non-first tiles get that TABLE's own column header prepended (not the
        top of the document), so Gemini always sees the correct column labels.

    Step 4 — Scale each tile independently:
        Tiles are scaled to MAX_PX on their longest side only if needed.

    Order: RTL (rightmost column group first) so the first tiles seen by the
    merge contain the م (serial) column — the deduplicator's primary key.
    """
    w, h   = img.size
    n_cols = max(1, round(w / COL_SPLIT_PX))
    col_w  = w // n_cols

    # First pass: collect layout info for all column strips
    col_info = []
    for c in range(n_cols - 1, -1, -1):   # RTL
        x0    = c * col_w
        x1    = w if c == n_cols - 1 else (c + 1) * col_w
        strip = img.crop((x0, 0, x1, h))
        first_header_y, cuts = _find_table_layout(strip)
        col_info.append((first_header_y, cuts, strip))

    # Cross-column consistency: if one column's first_header_y is far from the
    # consensus of others (> TILE_PX/2 away from median), it was likely triggered
    # by a decorative divider, not the actual column-header row.  Override it.
    detected_ys = [hy for hy, _, _ in col_info if hy > 0]
    if len(detected_ys) >= 2:
        median_y = sorted(detected_ys)[len(detected_ys) // 2]
        corrected = []
        for hy, cuts, strip in col_info:
            if hy > 0 and abs(hy - median_y) > TILE_PX // 2:
                log.info("Cross-col override: first_header_y %d → %d (median)", hy, median_y)
                # Also drop any boundary that no longer makes sense after the shift
                cuts = [b for b in cuts if b > median_y]
                hy   = median_y
            corrected.append((hy, cuts, strip))
        col_info = corrected

    all_tiles = []
    for first_header_y, cuts, strip in col_info:
        for tile in _horizontal_tiles(strip, forced_cuts=cuts,
                                      first_header_y=first_header_y):
            all_tiles.append(_enhance(_scale_tile(tile)))

    log.info("Smart tiles: %dx%d → %d col-group(s) → %d tiles total",
             w, h, n_cols, len(all_tiles))
    return all_tiles

def _compress_image(data: bytes, mime: str) -> list:
    """Prepare a non-GIF image for the API using the smart tiling pipeline."""
    if mime == "application/pdf":
        return [(data, mime)]
    from PIL import Image
    img = Image.open(BytesIO(data)).convert("RGB")
    result = [(_to_jpeg(t), "image/jpeg") for t in _smart_tiles(img)]
    log.info("Compressed %s → %d tile(s), JPEG sizes: %s KB",
             mime, len(result), [round(len(r[0])/1024) for r in result])
    return result

def _extract_gif_frames(data: bytes) -> list:
    """
    Extract deduplicated frames from a GIF and apply smart tiling.
    LOAD_TRUNCATED_IMAGES is set globally above so broken/large GIF streams
    are handled gracefully (filled with grey) instead of raising an exception.
    """
    from PIL import Image
    img = Image.open(BytesIO(data))
    n    = getattr(img, "n_frames", 1)
    seen, output = set(), []

    for i in range(n):
        try:
            img.seek(i)
            img.load()             # force full frame decode now, before any crop
            # Convert to RGB using the GIF's own palette first — this preserves
            # document colours (black text on white background).  Then materialise
            # the result into fully independent in-memory bytes so that downstream
            # crop/enhance/JPEG-save operations have no dependency on the GIF decoder
            # and cannot trigger "broken data stream" / "suspension not allowed here".
            rgb   = img.convert("RGB")   # uses actual GIF palette → correct colours
            raw   = rgb.tobytes()        # raw RGB bytes (3 per pixel, palette-free)
            frame = Image.frombytes("RGB", img.size, raw)
        except Exception as e:
            log.warning("GIF frame %d decode error (skipped): %s", i, e)
            continue

        # Dedup via 16×16 thumbnail hash
        thumb = frame.copy(); thumb.thumbnail((16, 16))
        fhash = hashlib.md5(thumb.tobytes()).hexdigest()
        if fhash in seen:
            _stats["frames_skipped"] += 1
            continue
        seen.add(fhash)

        for tile in _smart_tiles(frame):
            output.append((_to_jpeg(tile), "image/jpeg"))

    log.info("GIF: %d frame(s) → %d tiles (skipped %d dup frames)",
             n, len(output), _stats["frames_skipped"])
    return output

PROMPT = """You are a forensic Arabic table extractor for Gulf-region vehicle seizure/impound records (Dubai Police, Sharjah Police, Ajman Security, Emirates Auction). Your output is used for legal vehicle processing — a wrong VIN or plate digit causes real-world harm.

CORE PRINCIPLE — READ, DON'T GUESS:
1. Transcribe ONLY what is visually present. Blur, cut-off, or ambiguity → flag it, never guess.
2. Never pattern-complete. 14 visible VIN chars → return 14, not 17.
3. Never copy a value from an adjacent row. Each row is independent.
4. Masked values (*** or xxxx) → reproduce the mask EXACTLY. Never fill them in.
5. An empty cell is data. Return "" — do not borrow from anywhere.
6. NEVER invent rows to fill gaps in م (serial numbers). If the document shows م=1, 2, 5, 6 with rows 3 and 4 physically absent, your output must contain exactly those 4 rows (م values 1, 2, 5, 6). Missing serial numbers are real document features — do NOT synthesize the missing rows or carry over values from neighboring rows to fill them.

DOCUMENT LAYOUTS — detect before extracting:
• Layout A — single table, 30–40 rows (Ajman Security). Highest accuracy possible.
• Layout B — 3–4 side-by-side column-groups, 100+ rows, tiny text (Dubai Police, Emirates Auction). Default VIN confidence: medium.
• Layout C — 3 column-groups, 400+ rows, heavy compression (Sharjah Police). Default VIN confidence: low.
• Layout D — 2 column-groups, mixed full/masked VINs (Ajman Municipality).
Arabic RTL multi-group: rightmost column-group = group 1. Process right-to-left across groups.

PRIMARY COLUMNS — always extract using EXACTLY these Arabic names:
  م              → row serial number
  اللوحة         → vehicle LICENSE PLATE code (e.g. "14789", "بدون لوحة")
  مصدر اللوحة   → plate source emirate (e.g. "دبي", "أبوظبي", "الشارقة", "عجمان")
  فئة اللوحة    → plate category code (e.g. "1", "A", "الأولى")
  نوع المركبة   → vehicle make/model in Arabic
  لونها          → vehicle colour in Arabic
  تاريخ الحجز   → seizure date — normalize to YYYY-MM-DD (Gulf = DAY/MONTH/YEAR)
  الملاحظات     → VIN / chassis number (10–22 alphanumeric with Latin letters)

COLUMN ALIASES — some documents use different header names for the same canonical field:
  • رقم القاعدة  →  الملاحظات  (base/fleet number = chassis/VIN — confirmed by JTDBR22E… style values)
  • رقم المركبة  →  اللوحة     (vehicle registration number = plate number)
  • شاصية / رقم الشاصية / رقم الهيكل  →  الملاحظات  (all chassis/VIN aliases)
  • جهة الإصدار / الجهة المصدرة  →  مصدر اللوحة  (issuing authority = plate source emirate)
  Always use the canonical field name in your output headers — not the document's alias.

PLATE ATTRIBUTES vs PLATE NUMBER — critical distinction:
  اللوحة (plate number) = the digits printed on the plate, e.g. "14789". Always a short numeric string.
  The following describe the PLATE ITSELF — they are NOT the plate number, map them to فئة اللوحة:
  • كود اللوحة   = plate category code (a letter or small number: "A", "ب", "1", "الأولى")
  • لون اللوحة   = plate colour scheme (e.g. "أخضر وأبيض") — this encodes the plate TYPE, not the vehicle body colour
  • نوع اللوحة / تصنيف اللوحة = same plate-category concept
  Important: لون اللوحة ≠ لونها.  لونها = vehicle BODY colour ("أبيض", "أسود"). They are two separate columns.

COLUMNS THAT ARE NOT اللوحة — never put these into the plate number column:
  • كود اللوحة / لون اللوحة / فئة اللوحة  — plate attributes, map to فئة اللوحة.
  • الكود  — internal facility ID (sequential integer e.g. 8398). Keep as its own "الكود" column.
  • رقم التخزين  — impound storage/lot number. Keep as its own "رقم التخزين" column.
  • رقم القاعدة  — chassis/VIN (maps to الملاحظات, NOT to اللوحة).
  If the actual plate number column is رقم المركبة, map it to اللوحة and keep الكود separate.

EXTRA COLUMNS — if the document contains columns with no canonical equivalent (e.g. رقم القضية, تصنيف المركبة), include them AFTER the canonical columns using the EXACT header text from the document.

FIELD RULES:

VIN / الملاحظات (HIGHEST RISK):
- Also appears in documents as: رقم القاعدة, شاصية, رقم الشاصية, رقم الهيكل — treat all as الملاحظات.
- Standard VIN = 17 chars, A–Z + 0–9, NEVER I/O/Q.
- Transcribe char by char. Use VIN knowledge ONLY to FLAG — never silently correct.
  • If you read an "O": report "O" AND flag invalid_vin_char.
  • If you read 15 chars: return 15 AND flag wrong_length.
- Short non-standard formats (MA31-018527, 116032-12009804) → keep exactly, flag non_standard_format.
- Masked (LNYADDA***KN01900*) → reproduce asterisks, flag masked_in_source. NEVER fill the mask.
- Blank → "" — do NOT borrow from another row.
- OCR confusion pairs to flag (not auto-correct): O/0, I/1/l, Q/0, B/8, S/5, Z/2, G/6, D/0.

PLATE / اللوحة:
- Also appears in documents as: رقم المركبة — treat as اللوحة (it is the registration plate number).
- "بدون لوحة" or "بدون رقم" → return literal Arabic text, NEVER empty string.
- Partially obscured plate → flag plate_obscured.
- CRITICAL — three numeric columns that are NOT اللوحة:
  • الكود          = internal facility ID (sequential integer like 8398). Keep as its own "الكود" column. NEVER map to اللوحة.
  • رقم التخزين   = impound storage/lot number (often 7–10 digits). Keep as its own column. NEVER map to اللوحة.
  • رقم القاعدة   = chassis/VIN number (maps to الملاحظات). NEVER map to اللوحة.
  The column HEADER is the ONLY reliable way to tell them apart from a plate number.
  If the header row is not visible in this tile, DO NOT guess — flag the field as "column_ambiguous" instead.
  NEVER put a رقم التخزين or الكود or رقم القاعدة value into اللوحة — this identifies the wrong vehicle.

DATES / تاريخ الحجز:
- Gulf format = DAY/MONTH/YEAR. "8/3/2023" = 8 March → "2023-03-08".
- Store BOTH: normalized YYYY-MM-DD in "تاريخ الحجز" AND raw original in "تاريخ_الحجز_الأصلي".
- If both D and M are ≤ 12 (ambiguous): flag date_order_assumed.

UNREADABLE CHARACTERS — use the block symbol ■ to mark uncertainty at character level:
- Individual unreadable character (smudged / blurred / cut off but its POSITION is visible in the cell): replace that character with ■. Examples:
  • VIN where chars 5 and 9 are smudged → "1ABC■EFG■JK17"
  • Plate where last digit is unclear     → "1478■"
- Completely invisible cell (the cell exists but the ENTIRE content cannot be seen): return exactly "■■■" (three blocks, nothing else).
- Readable cells: use plain text only — do NOT add ■ to clear values.
- Do NOT guess: if a character could be 0 or O, pick the most likely AND flag it — do not silently replace it with ■.

CONFIDENCE LEVELS (apply to VIN, plate, plate_code, dates):
- "high" → every character is crisp and unambiguous
- "medium" → readable but ≥1 char sits in a confusion pair
- "low" → blurred, partially cut off, or genuinely uncertain
- "unreadable" → cannot read at all → return "■■■" + flag

FLAG TRIGGERS (set _needs_review: true on the field):
- VIN length ≠ 17 AND not a known non-standard format
- VIN contains I, O, or Q
- VIN/plate confidence is medium, low, or unreadable
- Date could not be parsed or day-first order was assumed
- Value is masked in source (***)
- Any character in a critical field was a confusion-pair coin-flip

OUTPUT SCHEMA — a single JSON object (NOT an array):
{
  "layout_detected": "A",
  "column_group_count": 1,
  "extraction_notes": "page-level observations: skew, cut-off rows, resolution issues",
  "headers": ["م","اللوحة","مصدر اللوحة","فئة اللوحة","نوع المركبة","لونها","تاريخ الحجز","الملاحظات","تاريخ_الحجز_الأصلي"],
  "rows": [
    {
      "م": "1",
      "اللوحة": "14789",
      "مصدر اللوحة": "الشارقة",
      "فئة اللوحة": "الأولى",
      "نوع المركبة": "سوزوكي سويفت",
      "لونها": "أبيض",
      "تاريخ الحجز": "2023-03-08",
      "الملاحظات": "JS2ZC21S885407577",
      "تاريخ_الحجز_الأصلي": "8/3/2023",
      "_cell_confidence": {"الملاحظات": "high", "اللوحة": "high"},
      "_needs_review": {},
      "_review_reasons": {}
    },
    {
      "م": "37",
      "اللوحة": "بدون لوحة",
      "الملاحظات": "LNYADDA***KN01900*",
      "_cell_confidence": {"الملاحظات": "low"},
      "_needs_review": {"الملاحظات": true},
      "_review_reasons": {"الملاحظات": "masked_in_source — contains ***, do not auto-complete"}
    }
  ]
}

TILING NOTE — large images are automatically sliced into tiles before being sent to you:
- Each tile may show only a vertical strip (one column group) of the original document.
- Extract ONLY the columns you can actually see in this tile — do NOT invent columns for other parts of the document.
- Tiles that show only 2-3 columns should report headers and rows for those 2-3 columns only.
- The system merges all tiles after the fact, so incomplete rows are fine — they will be filled in from other tiles.
- HEADER ROW: even if a tile starts mid-page, the document's column headers are shown at the very top of the tile image. Use them to identify every column before reading any values. If you cannot read a header clearly, flag that column as "column_ambiguous" — do not guess from the values alone.
- ROW GAPS: if a tile shows م=1,2,7,8 with nothing printed for 3–6, return exactly 4 rows. Do not fill in rows 3–6.

MULTI-TABLE BOUNDARY — CRITICAL:
Some tiles contain the end of one table AND the beginning of a second table separated by a visible header row or blank divider.
Rules:
1. Extract ONLY the FIRST table you encounter. Stop at the separator/second header. Do NOT continue into the second table.
2. A second table starts when you see a NEW header row (dark background row with column names) appearing mid-tile AFTER data rows.
3. When م resets back to 1 (or a small number far below the last row's م), that is the start of a new table — STOP there.
4. Do NOT repeat the last few rows of the first table to "fill" the gap before the second table starts.
5. Set "extraction_notes" to "multi_table_detected — extracted first table only, second table begins at row [م value]".
Violation of this rule causes the most severe form of hallucination — fabricated rows that don't exist in either table.

The goal: right about what is certain, honest about what is not."""

# ── Extraction Logic ──────────────────────────────────────────────────────────

LARGE_FILE_THRESHOLD = 3 * 1024 * 1024   # 3 MB — above this, use the Files API

def _upload_file(data: bytes, mime: str, name: str) -> str:
    """Upload to Gemini Files API using resumable protocol. Returns the file URI."""
    base_url = f"https://generativelanguage.googleapis.com/upload/v1beta/files?key={API_KEY}"
    # Step 1: initiate resumable session
    init_req = urllib.request.Request(
        base_url,
        data=json.dumps({"file": {"display_name": name}}).encode(),
        headers={
            "Content-Type": "application/json",
            "X-Goog-Upload-Protocol": "resumable",
            "X-Goog-Upload-Command": "start",
            "X-Goog-Upload-Header-Content-Length": str(len(data)),
            "X-Goog-Upload-Header-Content-Type": mime,
        },
        method="POST",
    )
    with urllib.request.urlopen(init_req, timeout=30) as r:
        upload_url = r.headers.get("X-Goog-Upload-URL")
    # Step 2: upload the bytes
    upload_req = urllib.request.Request(
        upload_url, data=data,
        headers={
            "Content-Length": str(len(data)),
            "X-Goog-Upload-Offset": "0",
            "X-Goog-Upload-Command": "upload, finalize",
        },
        method="POST",
    )
    with urllib.request.urlopen(upload_req, timeout=300) as r:
        return json.loads(r.read().decode())["file"]["uri"]

def _get_column_samples(ws, headers: list, n: int = 4) -> dict:
    """Read up to n non-empty sample values from each column of the worksheet."""
    samples = {}
    for col_idx, h in enumerate(headers, 1):
        if not h:
            continue
        vals = []
        for row in range(2, min(ws.max_row + 1, 200)):
            v = ws.cell(row=row, column=col_idx).value
            s = str(v).strip() if v is not None else ""
            if s:
                vals.append(s)
            if len(vals) >= n:
                break
        samples[str(h)] = vals
    return samples

def call_gemini(data: bytes, mime: str, name: str = "file",
                known_headers: list = None, column_samples: dict = None) -> str:
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL}:generateContent?key={API_KEY}"

    prompt = PROMPT
    if known_headers:
        prompt += (
            f"\n\nCOLUMN CONTRACT — this table has EXACTLY {len(known_headers)} columns in this fixed order:\n"
            f"{known_headers}\n"
            f"Every row MUST contain exactly {len(known_headers)} values in that exact order. "
            "Use \"\" for blank or illegible cells. NEVER add, remove, or reorder columns.\n\n"
            "COLUMN ACCURACY — each column has a specific data pattern. "
            "Read the examples below before filling any row. "
            "If a value does not match a column's pattern, re-check the image — you are likely reading the wrong cell:\n"
        )
        for h in known_headers:
            s = (column_samples or {}).get(str(h), [])
            if s:
                examples = "  |  ".join(f'"{v}"' for v in s[:3])
                prompt += f'  • {h} → {examples}\n'
            else:
                prompt += f'  • {h} → (no examples yet — infer from column header)\n'
        prompt += (
            "\nDo NOT swap values between columns. "
            "Chassis/VIN numbers (رقم الشاصي/الملاحظات) are long Latin alphanumeric codes like 'JN6FE5459SX417956'. "
            "Plate numbers (اللوحة) are short numeric codes (4–6 digits). "
            "Serial (م) is a small integer 1–999. "
            "Source (مصدر اللوحة) is an emirate name like دبي, أبوظبي, الشارقة, عجمان. "
            "If unsure, leave the cell as \"\" rather than guessing.\n\n"
            "IMPORTANT: Even if the visible column headers in THIS image differ from the contract above, "
            "you MUST still output the contract column names and map each visible column to the closest contract column. "
            "Do NOT invent new column names.\n\n"
            "BILINGUAL TABLE REMINDER: If this page shows both Arabic and English column labels "
            "(e.g. 'لونها / Color', 'نوع المركبة / Model'), output the Arabic column name only. "
            "Every data row belongs to ONE set of columns only — do NOT produce duplicate rows "
            "or split data across Arabic and English column names."
        )

    if len(data) > LARGE_FILE_THRESHOLD:
        # Files API does not support image/gif — convert to PNG first
        if mime == "image/gif":
            log.info("Converting GIF to PNG for Files API: %s", name)
            try:
                from PIL import Image
                img = Image.open(BytesIO(data))
                img.seek(0)          # first frame of animated GIFs
                img = img.convert("RGB")
                buf = BytesIO()
                img.save(buf, format="PNG")
                data, mime = buf.getvalue(), "image/png"
                log.info("Converted to PNG: %d bytes", len(data))
            except Exception as ce:
                log.warning("GIF→PNG conversion failed (%s), sending inline instead", ce)
                media_part = {"inlineData": {"mimeType": "image/gif", "data": base64.b64encode(data).decode()}}
                timeout = 300
                goto_inline = True
            else:
                goto_inline = False
        else:
            goto_inline = False

        if not goto_inline:
            log.info("UPLOAD via Files API: %s (%d bytes, %s)", name, len(data), mime)
            file_uri = _upload_file(data, mime, name)
            media_part = {"fileData": {"mimeType": mime, "fileUri": file_uri}}
        timeout = 300
    else:
        log.info("INLINE request: %s (%d bytes, %s)", name, len(data), mime)
        media_part = {"inlineData": {"mimeType": mime, "data": base64.b64encode(data).decode()}}
        timeout = 300  # same as Files API — Gemini can be slow under load

    _rate_limit()
    _stats["api_calls"] += 1
    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}, media_part]}],
        "generationConfig": {
            "responseMimeType": "application/json",
            "maxOutputTokens": 8192,
            "temperature": 0,
        }
    }).encode()
    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"}, method="POST")
    # 503/429 backoff: 15 s → 30 → 60 → 90 → 120 → 150 → 180 (7 attempts, ~10 min total)
    BACKOFF = [15, 30, 60, 90, 120, 150, 180]
    MAX_ATTEMPTS = len(BACKOFF) + 1
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode()
                log.info("Gemini response received (%d chars)", len(raw))
                return raw
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            if e.code in (429, 503) and attempt < MAX_ATTEMPTS:
                wait = BACKOFF[attempt - 1]
                log.warning("Gemini %d on attempt %d/%d — retrying in %ds", e.code, attempt, MAX_ATTEMPTS, wait)
                time.sleep(wait)
                continue
            try:
                msg = json.loads(body)["error"]["message"]
            except Exception:
                msg = body[:300]
            log.error("Gemini error %d: %s", e.code, msg)
            if e.code in (429, 503):
                raise RuntimeError(f"Gemini is overloaded (503) — all {MAX_ATTEMPTS} retries exhausted. Please retry this file in a few minutes.")
            raise RuntimeError(f"Gemini {e.code}: {msg}")
        except (TimeoutError, urllib.error.URLError) as e:
            if attempt < MAX_ATTEMPTS:
                wait = BACKOFF[attempt - 1]
                log.warning("Network error on attempt %d/%d (%s) — retrying in %ds", attempt, MAX_ATTEMPTS, e, wait)
                time.sleep(wait)
            else:
                raise RuntimeError(f"Gemini unreachable after {MAX_ATTEMPTS} attempts — check your connection and try again")

def _normalize_table_format(table: dict) -> dict:
    """
    Convert the new dict-row schema (rows are dicts with _needs_review etc.)
    to the legacy array-row format that merge_and_sort expects.
    Extracts Gemini's per-cell confidence and needs_review into parallel lists.
    """
    headers  = table.get("headers", [])
    raw_rows = table.get("rows",    [])
    if not raw_rows or not isinstance(raw_rows[0], dict):
        return table   # already array format — nothing to do

    array_rows, gemini_flags = [], []
    for row_dict in raw_rows:
        nr   = {k: v for k, v in row_dict.get("_needs_review",   {}).items() if v}
        rr   = row_dict.get("_review_reasons", {})
        conf = row_dict.get("_cell_confidence", {})
        # Promote medium/low/unreadable confidence to _needs_review
        for field, level in conf.items():
            if level in ("medium", "low", "unreadable") and field not in nr:
                nr[field]  = True
                rr[field]  = rr.get(field, f"confidence_{level}")
        # Detect ■ markers written by Gemini for unreadable content
        for h in headers:
            raw_val = str(row_dict.get(h, "") or "").strip()
            if raw_val == "■■■":
                # Entirely unseen cell — flag as UNSEEN so the dashboard can colour it red
                nr[h]  = True
                rr[h]  = "UNSEEN"
            elif "■" in raw_val:
                # Partially readable — flag so the dashboard highlights the ■ chars
                if h not in nr:
                    nr[h] = True
                    rr[h] = rr.get(h, "PARTIAL_UNREADABLE")

        # Store reason text (not bool) so it flows through the whole pipeline
        flag_with_reasons = {field: rr.get(field, "Gemini flagged") for field in nr}
        gemini_flags.append(flag_with_reasons)
        # Build ordered array from header list
        array_rows.append([str(row_dict.get(h, "") or "").strip() for h in headers])

    return {
        "title":               table.get("title", ""),
        "headers":             headers,
        "rows":                array_rows,
        "_gemini_flags":       gemini_flags,
        "layout_detected":     table.get("layout_detected",  ""),
        "extraction_notes":    table.get("extraction_notes", ""),
    }


def parse_response(raw: str) -> list:
    data = json.loads(raw)
    candidate = data["candidates"][0]

    finish = candidate.get("finishReason", "")
    if finish == "MAX_TOKENS":
        log.warning("Gemini hit MAX_TOKENS — response was truncated. Attempting partial recovery.")

    text = candidate["content"]["parts"][0]["text"].strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as e:
        log.warning("JSON truncated at char %d — attempting partial row recovery", e.pos)
        return _recover_truncated_json(text)

    # Normalize to a list of table objects
    if isinstance(parsed, dict):
        # New single-object schema
        tables = [parsed]
    elif isinstance(parsed, list):
        tables = parsed
    else:
        tables = [{"headers": [], "rows": []}]

    return [_normalize_table_format(t) for t in tables]

def _recover_truncated_json(text: str) -> list:
    """Extract all complete rows from a JSON response that was cut off mid-output."""
    import re
    # Pull out every complete row array: ["val1","val2",...]
    # A complete row ends with ] followed by , or newline (not mid-string)
    rows = re.findall(r'\[(?:[^\[\]]|"(?:[^"\\]|\\.)*")*\]', text)
    if not rows:
        raise RuntimeError("Gemini response was truncated and no rows could be recovered. Try a smaller image.")

    # First match may be the headers array — detect by checking if it's inside "headers":
    headers = None
    header_match = re.search(r'"headers"\s*:\s*(\[[^\]]+\])', text)
    if header_match:
        try:
            headers = json.loads(header_match.group(1))
        except Exception:
            pass

    # All remaining arrays are data rows
    data_rows = []
    for r in rows:
        try:
            parsed_row = json.loads(r)
            if parsed_row != headers:
                data_rows.append(parsed_row)
        except Exception:
            continue

    if not headers or not data_rows:
        raise RuntimeError("Gemini response was truncated too early to recover any rows.")

    log.info("Partial recovery: %d rows salvaged from truncated response", len(data_rows))
    return [{"title": "", "headers": headers, "rows": data_rows, "_truncated": True}]

def _is_data_row_masquerading_as_headers(headers: list) -> bool:
    """
    Return True when Gemini put a data row (or blank row) in the headers field
    instead of column names. Two triggers:
      • Majority are blank strings → Gemini returned empty column names
      • Majority look like data values (numbers / dates / VINs)
    """
    import re as _re
    if not headers:
        return True

    # Blank headers — Gemini failed to name the columns
    blank = sum(1 for h in headers if not str(h or "").strip())
    if blank >= max(2, len(headers) // 2):
        return True

    data_like = 0
    for h in headers:
        h_str = str(h or "").strip()
        if not h_str:
            continue
        if h_str.replace("/", "").replace("-", "").isdigit() and len(h_str) <= 7:
            data_like += 1
        elif _re.search(r'\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}', h_str):
            data_like += 1
        elif len(h_str) >= 10 and any(c.isdigit() for c in h_str) and any(c.isalpha() for c in h_str):
            data_like += 1
    return data_like >= max(2, len(headers) // 2)


def _type_assign_row(values: list, canonical_headers: list) -> list:
    """
    Last-resort row recovery: assign each value to the best-fitting canonical
    column based purely on value type (VIN / date / plate / serial).
    Text values (color, vehicle, source) are placed left-to-right into
    remaining untyped columns.
    Returns a row list aligned to canonical_headers.
    """
    import re as _re
    row = [""] * len(canonical_headers)

    # Index the typed canonical slots
    def slot(type_name):
        for i, h in enumerate(canonical_headers):
            if _get_col_type(h) == type_name and not row[i]:
                return i
        return None

    # Fallback: الملاحظات often stores VINs even though its type is "any"
    def vin_slot():
        s = slot("vin")
        if s is not None:
            return s
        for i, h in enumerate(canonical_headers):
            if "ملاحظات" in str(h) and not row[i]:
                return i
        return None

    # Known Arabic color words (partial list sufficient for matching)
    _COLORS = {"أبيض","ابيض","أسود","اسود","أحمر","احمر","أزرق","ازرق","أخضر","اخضر",
               "أصفر","اصفر","رمادي","رصاصي","بني","ذهبي","فضي","بيج","بنفسجي","برتقالي",
               "كحلي","زيتي","وردي","white","black","red","blue","green","silver","grey","gold"}
    _EMIRATES = {
        "دبي", "Dubai",
        "أبوظبي", "ابوظبي", "أبو ظبي", "ابو ظبي", "Abu Dhabi",
        "الشارقة", "الشارقه", "Sharjah",
        "عجمان", "Ajman",
        "رأس الخيمة", "رأس الخيمه", "راس الخيمة", "راس الخيمه", "Ras Al Khaimah",
        "الفجيرة", "الفجيره", "Fujairah",
        "أم القيوين", "أم القيون", "ام القيوين", "Umm Al Quwain",
        "خورفكان", "دبا", "كلباء",  # sub-cities
    }

    text_queue = []
    for val in values:
        v = str(val or "").strip()
        if not v:
            continue
        # VIN: 10-22 alphanumeric chars with both digits and LATIN (ASCII) letters
        if (10 <= len(v) <= 22 and any(c.isdigit() for c in v)
                and any(c.isascii() and c.isalpha() for c in v)):
            idx = vin_slot()
            if idx is not None:
                row[idx] = v
        # Date
        elif _re.search(r'\d{1,2}[/\-]\d{1,2}[/\-]\d{1,4}', v):
            idx = slot("date")
            if idx is not None:
                row[idx] = v
        # Pure digits
        elif v.replace("-", "").isdigit():
            digits = v.replace("-", "")
            if len(digits) <= 3:
                idx = slot("serial")
                if idx is not None:
                    row[idx] = v
                    continue
            if 4 <= len(digits) <= 9:
                idx = slot("plate")
                if idx is not None:
                    row[idx] = v
        else:
            text_queue.append(v)

    # Place text values semantically:
    # 1. Known color → لونها  2. Known emirate → مصدر اللوحة
    # 3. Single letter/code → فئة اللوحة  4. Longer text → نوع المركبة
    def _place_text(v):
        if v.lower() in _COLORS or v in _COLORS:
            for i, h in enumerate(canonical_headers):
                if "لون" in str(h) and not row[i]:
                    row[i] = v; return
        _em_low = {e.lower() for e in _EMIRATES}
        if v.lower() in _em_low:
            for i, h in enumerate(canonical_headers):
                if _get_col_type(h) == "source" and not row[i]:
                    row[i] = v; return
        _CAT_CODES = {
            "الأول","الثاني","الثالث","الرابع","الخامس","السادس","السابع","الثامن","التاسع","العاشر",
            "الأولى","الثانية","الثالثة","الرابعة","الخامسة","السادسة","السابعة","الثامنة","التاسعة","العاشرة",
            "أول","ثاني","ثالث","رابع",
        }
        is_code = (len(v) <= 2 and v.isascii()) or v in _CAT_CODES
        if is_code:
            for i, h in enumerate(canonical_headers):
                if "فئ" in str(h) and not row[i]:
                    row[i] = v; return
        if len(v) >= 5:                           # long text → vehicle type
            for i, h in enumerate(canonical_headers):
                if "نوع" in str(h) or "مركب" in str(h):
                    if not row[i]:
                        row[i] = v; return
        # Fallback: first empty column that isn't the plate number
        for i, h in enumerate(canonical_headers):
            t = _get_col_type(h)
            if t in ("any", "source") and not row[i]:
                row[i] = v; return

    for v in text_queue:
        _place_text(v)

    return row


def _score_column_order(values: list, headers: list) -> int:
    """
    Score how well `values` fit positionally into `headers` using typed-column validation.
    Each validatable column that passes adds +1; each that fails adds -1.
    Returns -999 if column counts don't match.
    """
    if len(values) != len(headers):
        return -999
    score = 0
    for val, h in zip(values, headers):
        col_type = _get_col_type(h)
        if col_type == "any":
            continue
        if _cell_ok(str(val).strip(), col_type):
            score += 1
        else:
            score -= 1
    return score


def _recover_malformed_tables(malformed: list, valid: list) -> list:
    """
    Attempt to recover tables whose 'headers' field is a data row OR blank.

    Two cases:
    • Data-in-headers: find the valid table whose column order best matches
      the values via type-scoring, then promote the 'headers' values to row 0.
    • Blank-headers: table has empty header names but real data in 'rows'.
      Find the best-fitting valid table and reuse its header names.
    """
    recovered = []
    for m in malformed:
        raw_headers = m.get("headers", [])
        extra_rows = m.get("rows", [])
        all_blank = all(not str(h or "").strip() for h in raw_headers)

        # Canonical headers from any valid table (used for type-assign fallback)
        canonical = valid[0]["headers"] if valid else []

        if all_blank:
            # Blank headers — real data is in rows; use positional scoring on first row
            if not extra_rows:
                log.warning("Malformed tile: blank headers and no rows — skipped")
                continue

            probe = extra_rows[0]
            best_t, best_score = None, -999
            for t in valid:
                s = _score_column_order(probe, t["headers"])
                if s > best_score:
                    best_score, best_t = s, t

            if best_t is not None and best_score >= 0:
                log.info("Recovering blank-header tile ('%s' order, %d rows, score=%d)",
                         best_t["headers"][0] if best_t["headers"] else "?", len(extra_rows), best_score)
                recovered.append({
                    "title": m.get("title", ""),
                    "headers": list(best_t["headers"]),
                    "rows": [list(r) for r in extra_rows],
                })
            else:
                # Positional scoring failed — reconstruct each row by value type
                rebuilt = [_type_assign_row(r, canonical) for r in extra_rows]
                rebuilt = [r for r in rebuilt if any(v for v in r)]
                if rebuilt:
                    log.info("Blank-header tile recovered via value-type (%d rows)", len(rebuilt))
                    recovered.append({"title": m.get("title", ""), "headers": canonical, "rows": rebuilt})
                else:
                    log.warning("Blank-header tile unrecoverable — dropped")
        else:
            # Data-in-headers case: try each valid table's column order
            best_t, best_score = None, -999
            for t in valid:
                s = _score_column_order(raw_headers, t["headers"])
                if s > best_score:
                    best_score, best_t = s, t

            if best_t is not None and best_score >= 0:
                log.info("Recovering malformed tile as row 0 ('%s' order, score=%d, +%d extra rows)",
                         best_t["headers"][0] if best_t["headers"] else "?", best_score, len(extra_rows))
                recovered.append({
                    "title": m.get("title", ""),
                    "headers": list(best_t["headers"]),
                    "rows": [list(raw_headers)] + [list(r) for r in extra_rows],
                })
            else:
                # Positional scoring failed — reconstruct each row by value type
                all_data = [list(raw_headers)] + [list(r) for r in extra_rows]
                rebuilt = [_type_assign_row(row, canonical) for row in all_data]
                rebuilt = [r for r in rebuilt if any(v for v in r)]
                if rebuilt:
                    log.info("Malformed tile recovered via value-type (%d rows, score was %s)",
                             len(rebuilt), best_score)
                    recovered.append({"title": m.get("title", ""), "headers": canonical, "rows": rebuilt})
                else:
                    log.warning("Malformed tile unrecoverable (score=%s) — dropped", best_score)
    return recovered


def _most_arabic_table(tables: list) -> dict:
    """Return the best canonical table.

    Priority: (1) has both a VIN and a date column, (2) most headers, (3) Arabic fraction.
    This prevents a tile with longer Arabic verbose headers (but no VIN column) from
    being chosen over the tile that actually contains all 8 canonical columns.
    """
    def quality_score(t):
        hdrs = t.get("headers", [])
        if not hdrs:
            return (0, 0, 0)
        chars = "".join(str(h) for h in hdrs)
        ar_frac = sum(1 for c in chars if "؀" <= c <= "ۿ") / len(chars) if chars else 0
        types = [_get_col_type(h) for h in hdrs]
        has_vin  = int("vin"  in types)
        has_date = int("date" in types)
        return (has_vin + has_date, len(hdrs), ar_frac)
    return max(tables, key=quality_score)


def _make_row_key(row: list, col_map: dict, vin_col, plate_col, serial_col) -> str:
    """Return a stable dedup key for a row: plate → VIN → serial.

    Plate-first: when two tiles show the same row with different amounts of data
    (one tile captured plate+source, another captured vehicle+VIN), both will key
    by plate and merge correctly — regardless of whether VIN is present.
    VIN as fallback handles rows where plate is missing/unreadable.
    Serial is last resort only (tiles may reuse the same serial indices).
    """
    vin_v = plate_v = serial_v = ""
    for i, val in enumerate(row):
        s = str(val or "").strip()
        if not s:
            continue
        c = col_map.get(i)
        if c == vin_col and _cell_ok(s, "vin"):
            vin_v = s
        elif c == plate_col and _cell_ok(s, "plate"):
            plate_v = s
        elif c == serial_col:
            serial_v = s
    if plate_v:
        return f"p:{plate_v}"
    if vin_v:
        return f"v:{vin_v}"
    if serial_v and serial_v.isdigit():
        return f"s:{serial_v}"
    return ""   # no reliable key — caller assigns a unique fallback


def _col_by_col_merge(tables: list, canonical: list) -> dict:
    """Column-first merge.

    Instead of picking one tile as the canonical row source, every tile
    contributes to a keyed row dictionary one column at a time:

    1. Map each tile's column headers → canonical column name.
    2. For each row in each tile, extract a unique key (VIN > serial > plate).
    3. For each canonical column, record the first non-empty value seen for
       that key across all tiles — later tiles fill in gaps, never overwrite.
    4. Reconstruct rows from the keyed dictionary and sort by serial number.

    Result: zero duplicate rows, no column dropped, no canonical tile selection.
    """
    serial_col = next((h for h in canonical if _get_col_type(h) == "serial"), None)
    plate_col  = next((h for h in canonical if _get_col_type(h) == "plate"),  None)
    vin_col    = next((h for h in canonical if _get_col_type(h) == "vin"),    None)

    rows_store:   dict = {}   # key → {canonical_col: value}
    gemini_store: dict = {}   # key → {field: reason} from Gemini _needs_review
    key_order:    list = []   # insertion order to preserve document order
    no_key_seq:   int  = 0

    for t in tables:
        hdrs = t.get("headers", [])
        rows = t.get("rows",    [])

        # Build column mapping for this tile
        col_map: dict = {}
        for i, h in enumerate(hdrs):
            m = _fuzzy_match_header(h, canonical)
            if m:
                col_map[i] = m
            else:
                log.info("col-merge: tile header '%s' has no canonical match — ignored", h)

        for row_idx, row in enumerate(rows):
            key = _make_row_key(row, col_map, vin_col, plate_col, serial_col)
            if not key:
                no_key_seq += 1
                key = f"_:{no_key_seq}"

            if key not in rows_store:
                rows_store[key] = {h: "" for h in canonical}
                gemini_store[key] = {}
                key_order.append(key)
            # Merge Gemini per-row flags (if this table carried them)
            gf_list = t.get("_gemini_flags", [])
            if gf_list and row_idx < len(gf_list):
                for f, v in gf_list[row_idx].items():
                    if v:
                        gemini_store[key][f] = v

            bucket = rows_store[key]
            for i, val in enumerate(row):
                c   = col_map.get(i)
                s   = str(val or "").strip()
                if c and s and not bucket[c]:   # fill empty slots only
                    bucket[c] = s

    # Reconstruct rows, drop fully-empty ones
    final_rows = []
    for k in key_order:
        row = [rows_store[k].get(h, "") for h in canonical]
        if any(str(v).strip() for v in row):
            final_rows.append(row)

    # Sort by serial number
    if serial_col and serial_col in canonical:
        si = canonical.index(serial_col)
        final_rows.sort(
            key=lambda r: (0, int(str(r[si]).strip()))
            if str(r[si]).strip().isdigit() else (1, 0)
        )

    log.info("col-merge: %d unique rows from %d tiles (canonical: %s)",
             len(final_rows), len(tables), canonical)

    # Build per-row Gemini flags aligned to final_rows order
    final_gemini_flags = [gemini_store.get(k, {}) for k in key_order
                          if any(rows_store[k].values())]

    return {
        "title":         tables[0].get("title", "Merged") if tables else "Merged",
        "headers":       canonical,
        "rows":          final_rows,
        "_gemini_flags": final_gemini_flags,
    }


def merge_and_sort(tables: list, canonical_override: list = None) -> dict:
    # Separate well-formed tables from malformed ones (data row in headers field)
    valid, malformed = [], []
    for t in tables:
        (malformed if _is_data_row_masquerading_as_headers(t.get("headers", [])) else valid).append(t)

    if not valid:
        valid = tables
    elif malformed:
        valid.extend(_recover_malformed_tables(malformed, valid))

    tables = valid

    # Determine canonical column list
    if canonical_override:
        canonical = list(canonical_override)
        log.info("merge: enforcing existing Excel canonical: %s", canonical)
    else:
        canonical = list(_most_arabic_table(tables)["headers"])

    return _col_by_col_merge(tables, canonical)

def _get_col_type(header: str) -> str:
    """Map a column header name to its expected value type."""
    h = _normalize(str(header or "")).lower()
    if any(k in h for k in ["شاصي", "هيكل", "ملاحظات", "chassis", "vin", "id"]):
        return "vin"
    if any(k in h for k in ["تاريخ", "date"]):
        return "date"
    if any(k in h for k in ["مصدر", "source"]):
        return "source"
    # "لوح" matches اللوحة correctly but also مصدر اللوحة / فئة اللوحة — exclude those
    # Also exclude storage/case/yard numbers that share the word "رقم"
    _plate_blocklist = ("تخزين", "مخزن", "مستودع", "موقع", "قضيه", "قضية", "ملف", "حادثه")
    if any(b in h for b in _plate_blocklist):
        return "any"
    if (any(k in h for k in ["رقم المركب", "plate"])
            or ("لوح" in h and "مصدر" not in h and "فئه" not in h)):
        return "plate"
    if any(k in h for k in ["متسلسل", "serial", " م "]) or h.strip() == "م":
        return "serial"
    return "any"

def _cell_ok(value, col_type: str) -> bool:
    """Return False if a cell value is clearly wrong for its column type."""
    import re
    v = str(value or "").strip()
    if not v:
        return True   # empty is always acceptable
    if col_type == "vin":
        # VIN/chassis: 10–22 chars, must contain both digits and LATIN (ASCII) letters
        return (10 <= len(v) <= 22
                and any(c.isdigit() for c in v)
                and any(c.isascii() and c.isalpha() for c in v))
    if col_type == "date":
        return bool(re.search(r'\d{1,4}[\/\-.]\d{1,2}[\/\-.]\d{1,4}', v))
    if col_type == "plate":
        return v.isdigit() and 2 <= len(v) <= 9
    if col_type == "serial":
        return v.replace("-", "").isdigit()
    return True

def _infer_dominant_type(values: list) -> str:
    """Return the dominant value type (vin/date/plate/serial) or 'any'."""
    typed = [str(v).strip() for v in values if str(v).strip()]
    if not typed:
        return "any"
    for ctype in ("vin", "date", "plate"):
        hits = sum(1 for v in typed if _cell_ok(v, ctype))
        if hits / len(typed) >= 0.7:
            return ctype
    return "any"


def _row_fingerprint(row_dict: dict, headers: list, vin_col: str = None) -> str:
    """
    Deduplication key for one row.
    VIN is used as the key when present (unique per vehicle, OCR-stable).
    Falls back to MD5 of all non-empty values — catches tile-overlap duplicates.
    """
    if vin_col:
        v = str(row_dict.get(vin_col) or "").strip().upper()
        if 10 <= len(v) <= 22:          # proper VIN/chassis length
            return "vin:" + v
    parts = [str(row_dict.get(h) or "").strip().lower() for h in headers]
    content = "|".join(p for p in parts if p)
    if not content:
        return ""                        # all-empty row — handled separately
    return "row:" + hashlib.md5(content.encode()).hexdigest()


def _resolve_headers(extracted_headers: list, extracted_rows: list,
                     existing_headers: list, existing_samples: dict = None) -> dict:
    """
    Map each extracted header to a canonical column name. Single-pass, no guessing.

    Returns dict: extracted_h → canonical_name
      - existing header name  → matched to an existing column
      - new string            → genuinely new column (will be created)
      - None                  → discard (blank header, all-empty column, or duplicate)
    """
    mapping = {}
    used = set()  # canonical names already claimed in this pass

    for idx, h in enumerate(extracted_headers):
        h_str = str(h or "").strip()

        if not h_str:
            mapping[h] = None
            continue

        # Does this column carry any real data?
        col_values = [row[idx] for row in extracted_rows if idx < len(row)]
        has_data = any(str(v).strip() for v in col_values)

        # Translate English header to Arabic before matching
        ar_translation = _EN_TO_AR.get(h_str.lower())
        search = ar_translation or h_str

        # Match against existing headers not yet claimed by another extracted column
        available = [e for e in existing_headers if e not in used]
        matched = _fuzzy_match_header(search, available)

        # Value-type cross-check: reject a fuzzy match where the column types contradict
        if matched and existing_samples and has_data:
            ex_type = _get_col_type(matched)
            if ex_type != "any":
                new_type = _infer_dominant_type(col_values)
                if new_type != "any" and new_type != ex_type:
                    log.info("Type mismatch: '%s'→'%s' rejected (%s vs %s)",
                             h, matched, new_type, ex_type)
                    matched = None

        # Value-type fallback: when name matching fails, route by what the values look like
        if not matched and has_data and existing_samples:
            new_type = _infer_dominant_type(col_values)
            if new_type != "any":
                for ex_h in available:
                    ex_samples_vals = [v for v in existing_samples.get(ex_h, []) if v]
                    if len(ex_samples_vals) < 2:
                        continue
                    ex_hits = sum(1 for v in ex_samples_vals if _cell_ok(v, new_type))
                    if ex_hits / len(ex_samples_vals) >= 0.7:
                        matched = ex_h
                        log.info("Value-type fallback: '%s' → '%s' (type=%s)", h, ex_h, new_type)
                        break

        if matched:
            mapping[h] = matched
            used.add(matched)
            if matched != h:
                log.info("Column mapped: '%s' → '%s'", h, matched)
        elif not has_data:
            # No match and no data — discard silently
            mapping[h] = None
            log.info("Column discarded (empty + unmatched): '%s'", h_str)
        else:
            # Genuinely new column with real data
            new_name = ar_translation or h_str
            # Reject if the new name would duplicate something already committed
            if new_name in used or _fuzzy_match_header(new_name, list(used)) is not None:
                log.warning("Column discarded (duplicate name): '%s' → '%s'", h_str, new_name)
                mapping[h] = None
            else:
                mapping[h] = new_name
                used.add(new_name)
                log.info("New column: '%s'", new_name)

    return mapping

def _normalize(s: str) -> str:
    """Canonical form for Arabic header comparison: remove diacritics, unify letter variants."""
    import unicodedata
    s = str(s or "").strip()
    # Remove diacritics (category Mn) and kashida
    s = "".join(c for c in s if unicodedata.category(c) != "Mn" and c != "ـ")
    # Unify Arabic letter variants so أ/إ/آ→ا, ة→ه, ى→ي, ؤ→و etc.
    for src, dst in [("أ","ا"),("إ","ا"),("آ","ا"),("ٱ","ا"),("ى","ي"),("ئ","ي"),("ة","ه"),("ؤ","و")]:
        s = s.replace(src, dst)
    return s.lower().strip()

# Known English↔Arabic header pairs for UAE vehicle/seizure tables.
# Add more as needed. Keys are lowercase English; values are canonical Arabic.
_EN_TO_AR: dict = {
    "serial no": "الرقم المتسلسل",
    "serial no.": "الرقم المتسلسل",
    "serial number": "الرقم المتسلسل",
    "no": "الرقم المتسلسل",
    "no.": "الرقم المتسلسل",
    "id": "رقم الشاصي",
    "vin": "رقم الشاصي",
    "chassis": "رقم الشاصي",
    "chassis no": "رقم الشاصي",
    "chassis no.": "رقم الشاصي",
    "chassis number": "رقم الشاصي",
    "plate": "رقم المركبة",
    "plate no": "رقم المركبة",
    "plate no.": "رقم المركبة",
    "plate number": "رقم المركبة",
    "vehicle no": "رقم المركبة",
    "vehicle no.": "رقم المركبة",
    "model": "نوع المركبة",
    "vehicle type": "نوع المركبة",
    "type": "نوع المركبة",
    "make": "نوع المركبة",
    "color": "لونها",
    "colour": "لونها",
    "date": "تاريخ الحجز",
    "seizure date": "تاريخ الحجز",
    "city": "مصدر اللوحة",
    "emirate": "مصدر اللوحة",
    "code": "فئة اللوحة",
    "code/qty": "فئة اللوحة",
    "category": "فئة اللوحة",
    "source": "مصدر اللوحة",
    "authority": "الجهة الحاجزة",
    "department": "الإدارة المركبة",
    "traffic fine": "التفليج المروري",
    "fine no": "رقم التفليج",
    "fine no.": "رقم التفليج",
    "notes": "ملاحظات",
    "remarks": "ملاحظات",
    "value": "العدد",
    "count": "العدد",
    "qty": "العدد",
    "quantity": "العدد",
    "record no": "رقم السجل",
    "record no.": "رقم السجل",
    "record number": "رقم السجل",
    "spec": "مواصفات المركبة",
    "specs": "مواصفات المركبة",
    "specifications": "مواصفات المركبة",
    "plate source": "مصدر اللوحة",
    "plate category": "فئة اللوحة",
}

# Semantic groups — any two headers that share a group are treated as the same column.
# Add variants here when a new document introduces a different spelling.
_SEMANTIC_GROUPS: list[set] = [
    # serial / sequence number
    {"م", "مسلسل", "الرقم المتسلسل", "الرقم التسلسلي", "رقم متسلسل", "رقم تسلسلي",
     "serial no", "serial no.", "serial number", "no", "no.", "#", "index"},
    # plate / vehicle number
    {"اللوحة", "رقم اللوحة", "رقم المركبة", "رقم اللوح", "الرقم",
     "plate", "plate no", "plate no.", "plate number", "vehicle no", "id/price"},
    # chassis / VIN — includes "الملاحظات" because UAE seizure docs store VINs there
    # رقم القاعدة (base/fleet number) appears in some documents as the VIN/chassis column
    {"رقم الشاصي", "رقم الهيكل", "شاصي", "هيكل", "شاصية", "رقم الشاصية",
     "الملاحظات", "رقم القاعدة", "القاعدة",
     "vin", "chassis", "chassis no", "chassis no.", "chassis number", "id",
     "reference id", "ref id", "ref. id", "vin/id"},
    # vehicle type / make / model
    {"نوع المركبة", "نوع السيارة", "المركبة", "النوع",
     "model", "vehicle type", "vehicle model", "type", "make",
     "vehicle make/model", "vehicle make", "make/model"},
    # vehicle colour (body paint) — لون اللوحة is plate-colour, not vehicle colour
    {"لونها", "لون المركبة", "لون السيارة", "color", "colour", "vehicle color", "vehicle colour"},
    # seizure / registration date — docs use both "حجز" and "تسجيل" for the date column
    {"تاريخ الحجز", "تاريخ التسجيل", "تاريخ الإصدار", "تاريخ الاصدار",
     "التاريخ", "تاريخ", "date", "seizure date", "registration date"},
    # plate source / emirate (issuance)
    # جهة الإصدار (issuing party) appears in some docs as the plate-source/emirate column
    {"مصدر اللوحة", "المصدر", "الإمارة", "الامارة", "المنطقة", "المدينة",
     "جهة الإصدار", "جهة الاصدار", "الجهة المصدرة",
     "source", "plate source", "city", "emirate", "region", "city/emirate"},
    # plate category / code / grade / colour-scheme
    # كود اللوحة / لون اللوحة both encode the plate's category (letter, number, or colour scheme)
    # NOT to be confused with: اللوحة (plate number) or لونها (vehicle body colour)
    {"فئة اللوحة", "الفئة", "فئة", "كود اللوحة", "لون اللوحة", "نوع اللوحة",
     "رمز اللوحة", "تصنيف اللوحة", "درجة اللوحة",
     "code", "code/qty", "category", "grade", "grade/type",
     "plate type", "plate category", "plate color", "plate colour", "plate code"},
    # emirate / city
    # notes / remarks (الملاحظات is in the chassis/VIN group — see above)
    {"ملاحظات", "notes", "remarks"},
    # internal vehicle code / registry ID — must NEVER be matched to اللوحة (plate)
    # الكود appears in some facility documents as an internal sequential vehicle ID
    {"الكود", "كود", "رمز المركبة", "الرمز",
     "vehicle code", "internal code", "code id"},
    # storage / yard / warehouse number — must NEVER match اللوحة (plate)
    {"رقم التخزين", "التخزين", "رقم المخزن", "رقم المستودع", "رقم الموقع",
     "storage no", "storage no.", "storage number", "yard no", "yard number", "lot no", "lot number"},
    # case / file number
    {"رقم القضية", "القضية", "رقم الملف", "رقم الحادثة",
     "case no", "case no.", "case number", "file no", "file number"},
    # location / depot
    {"الموقع", "المستودع", "المخزن", "location", "depot", "yard"},
    # authority / seizing party
    {"الجهة الحاجزة", "جهة الحجز", "authority", "seizing authority"},
    # admin department
    {"الإدارة المركبة", "الادارة", "department", "admin"},
    # value / count / amount
    {"العدد", "القيمة", "المبلغ", "value", "count", "qty", "quantity"},
    # record number
    {"رقم السجل", "السجل", "record no", "record no.", "record number"},
    # status / condition
    {"الحالة", "حالة", "status", "condition"},
]

# Pre-built normalized lookup: norm(alias) → index in _SEMANTIC_GROUPS
_NORM_GROUP_INDEX: dict[str, int] = {}
for _gi, _grp in enumerate(_SEMANTIC_GROUPS):
    for _alias in _grp:
        _NORM_GROUP_INDEX[_alias.strip().lower()] = _gi


def _fuzzy_match_header(new_h: str, existing: list) -> str | None:
    """
    Return the matching header from `existing` if new_h is close enough, else None.
    Priority: exact → semantic group → English translation → normalized → substring → similarity ≥ 0.82.
    """
    from difflib import SequenceMatcher

    # Step 0: exact match
    for ex in existing:
        if new_h == ex:
            return ex

    # Step 1: semantic group — any alias in the same group matches any existing alias
    key = new_h.strip().lower()
    gi = _NORM_GROUP_INDEX.get(key)
    if gi is None:
        gi = _NORM_GROUP_INDEX.get(_normalize(new_h).strip())
    if gi is not None:
        grp = {a.strip().lower() for a in _SEMANTIC_GROUPS[gi]}
        for ex in existing:
            if ex and ex.strip().lower() in grp:
                log.info("Semantic group match: '%s' → '%s'", new_h, ex)
                return ex
            if ex and _normalize(str(ex)).strip() in {_normalize(a) for a in _SEMANTIC_GROUPS[gi]}:
                log.info("Semantic group match (normalized): '%s' → '%s'", new_h, ex)
                return ex

    # Step 2: English header → known Arabic translation, then match against existing
    en_key = new_h.strip().lower()
    ar_translation = _EN_TO_AR.get(en_key)
    if ar_translation:
        n_trans = _normalize(ar_translation)
        for ex in existing:
            if ex and _normalize(str(ex)) == n_trans:
                log.info("EN→AR header match: '%s' → '%s'", new_h, ex)
                return ex
        # translation known but not yet in existing — signal caller to use the Arabic name
        log.info("EN→AR translation: '%s' → '%s' (not yet in sheet)", new_h, ar_translation)
        return None  # will be added with Arabic name via caller

    # Step 3: normalized + substring + similarity
    n_new = _normalize(new_h)
    best_score, best = 0.0, None
    for ex in existing:
        if ex is None:
            continue
        n_ex = _normalize(str(ex))
        if n_new == n_ex:
            return ex
        # Substring match only when both strings are long enough —
        # prevents single-char "م" from matching any word that contains م
        if n_new and n_ex and len(n_new) >= 2 and len(n_ex) >= 2 and (n_new in n_ex or n_ex in n_new):
            return ex
        score = SequenceMatcher(None, n_new, n_ex).ratio()
        if score > best_score:
            best_score, best = score, ex
    if best_score >= 0.82:
        log.info("Fuzzy header match: '%s' → '%s' (score=%.2f)", new_h, best, best_score)
        return best
    return None

def _style_data_sheet(ws):
    """Apply dark-blue header row, freeze pane, and auto-filter to a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    ws.row_dimensions[1].height = 20

def _update_meta_sheet(wb, source_file: str, rows_added: int):
    """Maintain a hidden _Meta sheet tracking every extraction run."""
    name = "_Meta"
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(["source_file", "extracted_at", "rows_added"])
        ws.sheet_state = "hidden"
    else:
        ws = wb[name]
    ws.append([source_file, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), rows_added])

def build_excel(table: dict, base_bytes: bytes = None, source_file: str = "") -> tuple:
    """
    Append extracted table to base_bytes (existing Excel) or create fresh.
    Returns (excel_bytes, rows_added, total_rows).
    """
    import pandas as pd
    from openpyxl import load_workbook

    new_df = pd.DataFrame(table["rows"], columns=table["headers"])

    if base_bytes:
        wb = load_workbook(BytesIO(base_bytes))
    else:
        wb = None

    if wb:
        ws = wb.active
        existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        existing_samples = _get_column_samples(ws, existing_headers)
        header_map = _resolve_headers(table["headers"], table["rows"], existing_headers, existing_samples)

        # Schema is locked once the sheet has data — never add new columns.
        # Any extracted column that didn't map to an existing header is dropped.
        new_cols = [canon for _, canon in header_map.items()
                    if canon is not None and canon not in existing_headers]
        seen = set()
        new_cols = [c for c in new_cols if not (c in seen or seen.add(c))]
        if new_cols:
            log.warning("Schema is locked — dropping %d unrecognised column(s): %s", len(new_cols), new_cols)
            # Null out the mapping for any header that resolved to a non-existing column
            header_map = {h: (v if v in existing_headers else None) for h, v in header_map.items()}

        # Pre-compute column types for suspicious-cell highlighting
        from openpyxl.styles import PatternFill
        _yellow = PatternFill("solid", fgColor="FFEB3B")
        col_types = {h: _get_col_type(h) for h in existing_headers}

        # Build fingerprint set from every existing data row for deduplication
        vin_col = next((h for h in existing_headers if _get_col_type(h) == "vin"), None)
        seen_fps: set = set()
        for r_idx in range(2, ws.max_row + 1):
            rv = {h: str(ws.cell(r_idx, c).value or "").strip()
                  for c, h in enumerate(existing_headers, 1)}
            fp = _row_fingerprint(rv, existing_headers, vin_col)
            if fp:
                seen_fps.add(fp)

        rows_before = ws.max_row
        skip_empty = skip_dup = 0
        for row in table["rows"]:
            rd = {}
            for h, val in zip(table["headers"], row):
                canon = header_map.get(h)
                if canon:
                    rd[canon] = val
            excel_row = [rd.get(h, "") for h in existing_headers]

            # Skip all-empty rows
            if not any(str(v).strip() for v in excel_row):
                skip_empty += 1
                continue

            # Skip duplicate rows (tile overlap or re-upload)
            fp = _row_fingerprint(rd, existing_headers, vin_col)
            if fp and fp in seen_fps:
                skip_dup += 1
                continue
            if fp:
                seen_fps.add(fp)

            ws.append(excel_row)

            # Highlight cells whose value doesn't match the column's expected type
            current_row = ws.max_row
            for col_idx, (h, val) in enumerate(zip(existing_headers, excel_row), 1):
                ct = col_types.get(h, "any")
                if ct != "any" and not _cell_ok(val, ct):
                    ws.cell(row=current_row, column=col_idx).fill = _yellow

        if skip_empty or skip_dup:
            log.info("Skipped: %d empty row(s), %d duplicate row(s)", skip_empty, skip_dup)
        rows_added = ws.max_row - rows_before
        total_rows = ws.max_row - 1
    else:
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            new_df.to_excel(writer, index=False, sheet_name="Data")
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        rows_added = len(table["rows"])
        total_rows = len(table["rows"])

    # Style + auto-width
    _style_data_sheet(ws)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    _update_meta_sheet(wb, source_file, rows_added)

    out = BytesIO()
    wb.save(out)
    excel_bytes = out.getvalue()

    if IS_LOCAL:
        try:
            LOCAL_EXCEL.write_bytes(excel_bytes)
        except PermissionError:
            raise RuntimeError("Excel file is open — please close it and try again.")

    return excel_bytes, rows_added, total_rows

# ── Dashboard Excel Sync ──────────────────────────────────────────────────────
STORE_EXCEL = DATA_DIR / "extracted_tables.xlsx"
_CANONICAL_COLS = ["م", "اللوحة", "مصدر اللوحة", "فئة اللوحة",
                   "نوع المركبة", "لونها", "تاريخ الحجز", "الملاحظات"]

def sync_excel_store(store: dict):
    """Rebuild the dashboard Excel from the current row_store state."""
    rows = store.get("rows", [])
    if not rows:
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment as _Comment

    # Discover extra columns: any non-private key not in canonical set, sorted for stability
    _fixed = {"source_file", "review_status"} | set(_CANONICAL_COLS)
    _seen_extra: list = []
    _seen_set: set = set()
    for _r in rows:
        for _k in _r:
            if not _k.startswith("_") and _k not in _fixed and _k not in _seen_set:
                _seen_extra.append(_k)
                _seen_set.add(_k)
    _seen_extra.sort()   # alphabetical so column order is stable across runs
    all_cols = ["source_file"] + _CANONICAL_COLS + _seen_extra + ["review_status"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    ws.sheet_view.rightToLeft = True

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin     = Side(style="thin", color="BDD7EE")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr      = Alignment(horizontal="center", vertical="center")

    for ci, col in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = ctr;  cell.border = bdr
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(rows)+1}"

    flag_fill     = PatternFill("solid", fgColor="FFE699")
    reviewed_fill = PatternFill("solid", fgColor="E2EFDA")
    alt_fill      = PatternFill("solid", fgColor="F0F6FF")
    norm_font     = Font(name="Calibri", size=10)
    flag_font     = Font(name="Calibri", size=10, color="7B3F00")

    for ri, row in enumerate(rows, 2):
        flags       = row.get("_flags", {})
        is_reviewed = row.get("_all_reviewed", False)
        for ci, col in enumerate(all_cols, 1):
            if col == "review_status":
                fc = len(flags)
                value = "✅ Clean" if fc == 0 else ("👁 Reviewed" if is_reviewed else f"⚠ {fc} flag(s)")
            else:
                value = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = bdr
            cell.alignment = Alignment(vertical="center")
            if col in flags and not is_reviewed:
                cell.fill = flag_fill
                cell.font = flag_font
                try:
                    cell.comment = _Comment(flags[col], "Table Extractor")
                except Exception:
                    pass
            elif is_reviewed:
                cell.fill = reviewed_fill; cell.font = norm_font
            elif ri % 2 == 0:
                cell.fill = alt_fill;      cell.font = norm_font
            else:
                cell.font = norm_font

    for ci, col in enumerate(all_cols, 1):
        vals = [str(row.get(col, "") or "") for row in rows]
        w = max([len(col)] + [len(v) for v in vals], default=8)
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 50)

    wb.save(str(STORE_EXCEL))

# ── Flask App ─────────────────────────────────────────────────────────────────

try:
    from flask import Flask, request, jsonify, make_response
except ImportError:
    import subprocess; subprocess.run([sys.executable, "-m", "pip", "install", "flask", "-q"], check=False)
    from flask import Flask, request, jsonify, make_response

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB

@app.route("/")
def index():
    return HTML

@app.route("/resume")
def resume():
    """Return the previously saved local Excel so the client can restore its session."""
    if IS_LOCAL and LOCAL_EXCEL.exists():
        from openpyxl import load_workbook as lw
        try:
            wb = lw(str(LOCAL_EXCEL), read_only=True, data_only=True)
            rows = max(wb.active.max_row - 1, 0)
            wb.close()
        except Exception:
            rows = 0
        resp = make_response(LOCAL_EXCEL.read_bytes())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["X-Total-Rows"] = str(rows)
        resp.headers["Access-Control-Expose-Headers"] = "X-Total-Rows"
        return resp
    return ("", 204)

@app.route("/clear", methods=["POST"])
def clear():
    """Delete the local backup file and wipe the cache so a fresh start is fully clean."""
    if IS_LOCAL and LOCAL_EXCEL.exists():
        try:
            LOCAL_EXCEL.unlink()
        except Exception:
            pass
    wiped = 0
    for f in CACHE_DIR.glob("*.json"):
        try:
            f.unlink()
            wiped += 1
        except Exception:
            pass
    if wiped:
        log.info("Cache cleared: %d file(s) deleted", wiped)
    return ("", 204)

@app.route("/config")
def config():
    return jsonify({"model": MODEL, "is_local": IS_LOCAL})

@app.route("/logs")
def logs():
    if not _log_file.exists():
        return ("No logs yet.", 200, {"Content-Type": "text/plain; charset=utf-8"})
    lines = _log_file.read_text(encoding="utf-8").splitlines()
    return ("\n".join(lines[-200:]), 200, {"Content-Type": "text/plain; charset=utf-8"})

@app.route("/stats")
def stats():
    return jsonify({
        **_stats,
        "cache_files": len(list(CACHE_DIR.glob("*.json"))),
        "model": MODEL,
    })

# ── Dashboard routes ───────────────────────────────────────────────────────────

@app.route("/dashboard")
def dashboard():
    tpl = Path(__file__).parent / "templates" / "dashboard.html"
    if not tpl.exists():
        return "Dashboard template not found. Create templates/dashboard.html.", 404
    return tpl.read_text(encoding="utf-8"), 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/api/rows")
def api_get_rows():
    return jsonify(load_row_store())

@app.route("/api/rows/<int:row_id>", methods=["PATCH"])
def api_patch_row(row_id):
    body  = request.get_json()
    store = load_row_store()
    row   = next((r for r in store["rows"] if r["_id"] == row_id), None)
    if not row:
        return jsonify({"error": "Row not found"}), 404
    field = body.get("field")
    if field and "value" in body:
        # Record the correction for the learning loop (before overwriting the old value)
        if body["value"] != row.get(field, ""):
            reason = (row.get("_flags", {}).get(field) or "manual_edit")
            try:
                record_correction(
                    source_file  = row.get("source_file", "unknown"),
                    field        = field,
                    model_read   = row.get(field, ""),
                    human_value  = body["value"],
                    reason       = reason,
                )
            except Exception as _ce:
                log.warning("record_correction failed (non-fatal): %s", _ce)
        row[field] = body["value"]
    if "reviewed" in body and field:
        row.setdefault("_flags", {})
        if body["reviewed"]:
            row["_flags"].pop(field, None)
        else:
            row["_flags"][field] = "Manually flagged"
    row["_all_reviewed"] = len(row.get("_flags", {})) == 0
    save_row_store(store)
    try:
        sync_excel_store(store)
    except Exception as e:
        log.warning("sync_excel_store failed: %s", e)
    return jsonify({"ok": True, "row": row})

@app.route("/api/rows/<int:row_id>", methods=["DELETE"])
def api_delete_row(row_id):
    store = load_row_store()
    before = len(store["rows"])
    store["rows"] = [r for r in store["rows"] if r["_id"] != row_id]
    if len(store["rows"]) == before:
        return jsonify({"error": "Row not found"}), 404
    save_row_store(store)
    try:
        sync_excel_store(store)
    except Exception as e:
        log.warning("sync_excel_store after delete failed: %s", e)
    return jsonify({"ok": True})

@app.route("/api/rows/bulk-delete", methods=["POST"])
def api_bulk_delete():
    ids = set(request.get_json().get("ids", []))
    if not ids:
        return jsonify({"error": "No ids provided"}), 400
    store = load_row_store()
    store["rows"] = [r for r in store["rows"] if r["_id"] not in ids]
    save_row_store(store)
    try:
        sync_excel_store(store)
    except Exception as e:
        log.warning("sync_excel_store after bulk delete failed: %s", e)
    return jsonify({"ok": True, "deleted": len(ids)})

@app.route("/api/export/excel")
def api_export_excel():
    from flask import send_file as _send_file
    store = load_row_store()
    sync_excel_store(store)
    if not STORE_EXCEL.exists():
        return jsonify({"error": "No data yet — upload files first"}), 404
    return _send_file(
        str(STORE_EXCEL), as_attachment=True,
        download_name="extracted_tables.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/api/stats")
def api_dashboard_stats():
    store    = load_row_store()
    rows     = store["rows"]
    flagged  = [r for r in rows if r.get("_flags")]
    reviewed = [r for r in rows if r.get("_all_reviewed")]
    return jsonify({
        "total_rows":   len(rows),
        "flagged_rows": len(flagged),
        "reviewed_rows": len(reviewed),
        "pending_rows": len([r for r in flagged if not r.get("_all_reviewed")]),
        "total_flags":  sum(len(r.get("_flags", {})) for r in rows),
        "source_files": list({r.get("source_file", "unknown") for r in rows}),
    })

@app.route("/upload", methods=["POST"])
def upload():
    doc  = request.files.get("file")   # image or PDF to extract
    base = request.files.get("base")   # accumulated Excel from client (optional)

    if not doc:
        return jsonify({"error": "No file received"}), 400

    fname = doc.filename.lower()
    if   fname.endswith((".jpg", ".jpeg")): mime = "image/jpeg"
    elif fname.endswith(".png"):             mime = "image/png"
    elif fname.endswith(".webp"):            mime = "image/webp"
    elif fname.endswith(".gif"):             mime = "image/gif"
    elif fname.endswith(".pdf"):             mime = "application/pdf"
    else:
        return jsonify({"error": "Unsupported format. Use JPG, PNG, WEBP, GIF, or PDF."}), 400

    log.info("--- Upload: %s (%s) ---", doc.filename, mime)
    try:
        raw_data   = doc.read()
        base_bytes = base.read() if base else None
        _stats["files"] += 1

        # Read headers + column samples from existing Excel so Gemini stays accurate
        known_headers = []
        column_samples = {}
        if base_bytes:
            try:
                from openpyxl import load_workbook as _lw
                _wb = _lw(BytesIO(base_bytes))          # not read_only — need cell() access for samples
                _ws = _wb.active
                known_headers = [_ws.cell(1, c).value for c in range(1, _ws.max_column + 1)
                                 if _ws.cell(1, c).value]
                column_samples = _get_column_samples(_ws, known_headers)
                _wb.close()
                log.info("Known headers: %s", known_headers)
                log.info("Column samples: %s", {k: v for k, v in column_samples.items() if v})
            except Exception as he:
                log.warning("Could not read existing headers/samples: %s", he)

        # Build tiles: GIFs expand to (frame×tile) list, other images to tile list
        if mime == "image/gif":
            tiles = _extract_gif_frames(raw_data)
        else:
            tiles = _compress_image(raw_data, mime)

        all_tables = []
        for i, (tile_data, tile_mime) in enumerate(tiles):
            cache_key = hashlib.md5(tile_data).hexdigest()
            cached = _cache_get(cache_key)
            if cached:
                log.info("Cache hit for tile %d of %s", i, doc.filename)
                _stats["cache_hits"] += 1
                all_tables.extend(cached)
                continue

            label  = doc.filename if len(tiles) == 1 else f"{doc.filename}[tile{i}]"
            raw    = call_gemini(tile_data, tile_mime, label, known_headers, column_samples)
            parsed = parse_response(raw)

            if any(t.get("_truncated") for t in parsed):
                log.warning("Tile %d was truncated — skipping cache so next upload re-extracts it", i)
            else:
                _cache_set(cache_key, parsed)
            all_tables.extend(parsed)

        tables = all_tables if all_tables else [{"headers": [], "rows": []}]

        for _ti, _t in enumerate(tables):
            log.info("Tile %d: %d rows | headers: %s", _ti, len(_t.get("rows", [])), _t.get("headers", []))
        log.info("Parsed %d table(s) total", len(tables))
        table      = merge_and_sort(tables, canonical_override=known_headers if known_headers else None)
        excel_bytes, rows_added, total_rows = build_excel(table, base_bytes, doc.filename)
        _stats["rows"] += rows_added
        try:
            merge_extracted_rows(table, doc.filename)
        except Exception as _mre:
            log.warning("Row store update failed (non-fatal): %s", _mre)
        log.info("Done: +%d rows, %d total | api_calls=%d cache_hits=%d",
                 rows_added, total_rows, _stats["api_calls"], _stats["cache_hits"])

        resp = make_response(excel_bytes)
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["X-Rows-Added"]  = str(rows_added)
        resp.headers["X-Total-Rows"]  = str(total_rows)
        resp.headers["Access-Control-Expose-Headers"] = "X-Rows-Added, X-Total-Rows"
        return resp

    except Exception as e:
        log.error("Upload failed for %s: %s", doc.filename, e, exc_info=True)
        return jsonify({"error": str(e)}), 500

# ── Embedded UI ───────────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Table Extractor</title>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Outfit:wght@700;800&family=Fira+Code:wght@400;500&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/@tabler/icons-webfont@3.19.0/dist/tabler-icons.min.css">
  <style>
    *, *::before, *::after { box-sizing:border-box; margin:0; padding:0; }
    :root {
      --bg:#09090b; --surface:#18181b; --surface2:#27272a;
      --border:#27272a; --border2:#3f3f46;
      --text:#f4f4f5; --muted:#71717a;
      --primary:#6366f1; --pdim:rgba(99,102,241,.15);
      --success:#10b981; --sdim:rgba(16,185,129,.12);
      --danger:#ef4444;  --ddim:rgba(239,68,68,.12);
      --r:12px; --rs:8px;
    }
    body { font-family:'Inter',sans-serif; background:var(--bg); color:var(--text); min-height:100vh; display:flex; flex-direction:column; }

    header { background:var(--surface); border-bottom:1px solid var(--border); padding:0 1.5rem; height:58px; display:flex; align-items:center; justify-content:space-between; position:sticky; top:0; z-index:50; }
    .brand { display:flex; align-items:center; gap:.6rem; text-decoration:none; }
    .brand-icon { width:34px; height:34px; border-radius:8px; background:var(--pdim); border:1px solid rgba(99,102,241,.25); display:flex; align-items:center; justify-content:center; color:var(--primary); font-size:1.1rem; }
    .brand-name { font-family:'Outfit',sans-serif; font-size:1.05rem; font-weight:700; color:var(--text); }
    .brand-name span { color:var(--primary); }
    .hdr-right { display:flex; align-items:center; gap:.6rem; }
    .row-badge { display:flex; align-items:center; gap:.4rem; background:var(--surface2); border:1px solid var(--border2); border-radius:50px; padding:.3rem .85rem; font-size:.78rem; font-weight:600; font-family:'Fira Code',monospace; }
    .dot { width:6px; height:6px; border-radius:50%; background:var(--success); box-shadow:0 0 6px var(--success); }
    .btn { display:inline-flex; align-items:center; gap:.4rem; padding:.45rem .95rem; border-radius:var(--rs); font-size:.82rem; font-weight:500; cursor:pointer; border:1px solid var(--border2); background:var(--surface2); color:var(--text); transition:all .15s; font-family:inherit; }
    .btn:hover { background:#3f3f46; }
    .btn:disabled { opacity:.45; cursor:not-allowed; }
    .btn-success { background:var(--success); border-color:var(--success); color:#fff; }
    .btn-success:hover:not(:disabled) { background:#059669; border-color:#059669; }
    .btn-primary { background:var(--primary); border-color:var(--primary); color:#fff; }
    .btn-primary:hover:not(:disabled) { background:#4f46e5; }

    main { flex:1; max-width:820px; width:100%; margin:0 auto; padding:1.75rem 1.5rem; display:flex; flex-direction:column; gap:1.2rem; }

    /* Session bar */
    .session-bar { background:var(--surface); border:1px solid var(--border); border-radius:var(--rs); padding:.65rem 1rem; display:flex; align-items:center; justify-content:space-between; gap:1rem; flex-wrap:wrap; }
    .session-left { display:flex; align-items:center; gap:.65rem; flex-wrap:wrap; }
    .session-status { display:flex; align-items:center; gap:.4rem; font-size:.8rem; color:var(--muted); }
    .session-status.has-data { color:var(--success); }
    .session-status i { font-size:1rem; }
    .session-right { display:flex; align-items:center; gap:.5rem; }

    /* Drop zone */
    .dropzone { background:var(--surface); border:2px dashed var(--border2); border-radius:var(--r); padding:3rem 2rem; text-align:center; cursor:pointer; transition:all .2s; position:relative; overflow:hidden; }
    .dropzone::after { content:''; position:absolute; inset:0; background:var(--pdim); opacity:0; transition:opacity .2s; pointer-events:none; }
    .dropzone:hover, .dropzone.over { border-color:var(--primary); box-shadow:0 0 28px rgba(99,102,241,.1); }
    .dropzone.over::after { opacity:1; }
    .dz-icon { width:70px; height:70px; border-radius:50%; background:var(--pdim); border:1px solid rgba(99,102,241,.2); display:flex; align-items:center; justify-content:center; font-size:2rem; color:var(--primary); margin:0 auto 1.1rem; position:relative; z-index:1; }
    .dz-title { font-family:'Outfit',sans-serif; font-size:1.25rem; font-weight:800; margin-bottom:.35rem; position:relative; z-index:1; }
    .dz-sub { color:var(--muted); font-size:.85rem; margin-bottom:1.1rem; position:relative; z-index:1; }
    .dz-tags { display:flex; align-items:center; justify-content:center; gap:.4rem; flex-wrap:wrap; position:relative; z-index:1; }
    .dz-tag { background:var(--surface2); border:1px solid var(--border2); border-radius:4px; padding:.18rem .48rem; font-size:.7rem; font-family:'Fira Code',monospace; color:var(--muted); }

    /* Queue */
    .queue-wrap { display:flex; flex-direction:column; gap:.65rem; }
    .queue-hdr { display:flex; align-items:center; justify-content:space-between; }
    .queue-label { font-size:.75rem; font-weight:600; text-transform:uppercase; letter-spacing:.06em; color:var(--muted); }
    .clear-btn { background:none; border:none; color:var(--muted); font-size:.75rem; font-family:inherit; cursor:pointer; display:flex; align-items:center; gap:.3rem; padding:.2rem .4rem; border-radius:4px; transition:all .15s; }
    .clear-btn:hover { color:var(--danger); background:var(--ddim); }
    .queue-list { display:flex; flex-direction:column; gap:.45rem; }
    .qi { background:var(--surface); border:1px solid var(--border); border-radius:var(--rs); padding:.8rem 1rem; display:flex; align-items:center; gap:.85rem; transition:border-color .2s; }
    .qi.processing { border-color:var(--primary); }
    .qi.done  { border-color:rgba(16,185,129,.35); }
    .qi.error { border-color:rgba(239,68,68,.35); }
    .qi-ico { width:36px; height:36px; border-radius:8px; display:flex; align-items:center; justify-content:center; font-size:1.05rem; flex-shrink:0; }
    .qi-ico.img { background:var(--pdim); color:var(--primary); }
    .qi-ico.pdf { background:var(--ddim); color:var(--danger); }
    .qi-body { flex:1; min-width:0; }
    .qi-name { font-size:.855rem; font-weight:500; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; margin-bottom:.18rem; }
    .qi-meta { font-size:.73rem; color:var(--muted); font-family:'Fira Code',monospace; }
    .qi-stat { flex-shrink:0; display:flex; align-items:center; gap:.4rem; font-size:.8rem; font-weight:500; white-space:nowrap; }
    .qi-stat.pending    { color:var(--muted); }
    .qi-stat.processing { color:var(--primary); }
    .qi-stat.done       { color:var(--success); }
    .qi-stat.error      { color:var(--danger); }

    @keyframes spin { to { transform:rotate(360deg); } }
    .spinner { width:15px; height:15px; border:2px solid rgba(99,102,241,.3); border-top-color:var(--primary); border-radius:50%; animation:spin .75s linear infinite; }

    .toast { position:fixed; bottom:1.25rem; right:1.25rem; background:var(--surface); border:1px solid var(--border2); border-radius:var(--rs); padding:.7rem 1rem; font-size:.83rem; display:flex; align-items:center; gap:.5rem; box-shadow:0 8px 24px rgba(0,0,0,.5); transform:translateY(80px); opacity:0; transition:all .3s cubic-bezier(.16,1,.3,1); z-index:200; max-width:360px; }
    .toast.show { transform:translateY(0); opacity:1; }
    .toast.s { border-color:rgba(16,185,129,.4); }
    .toast.e { border-color:rgba(239,68,68,.4); }
    .t-s { color:var(--success); }
    .t-e { color:var(--danger); }
  </style>
</head>
<body>

<header>
  <a class="brand" href="#">
    <div class="brand-icon"><i class="ti ti-table-shortcut"></i></div>
    <div class="brand-name">Table<span>Extractor</span></div>
  </a>
  <div class="hdr-right">
    <div class="row-badge"><span class="dot"></span><span id="total">0 rows</span></div>
    <a class="btn btn-primary" href="/dashboard" style="text-decoration:none;">
      <i class="ti ti-layout-dashboard"></i> Open Dashboard
    </a>
    <button class="btn btn-success" id="dl-btn" disabled>
      <i class="ti ti-download"></i> Download Excel
    </button>
  </div>
</header>

<main>

  <!-- Session bar -->
  <div class="session-bar">
    <div class="session-left">
      <div class="session-status" id="session-status">
        <i class="ti ti-info-circle"></i>
        <span id="session-label">No data yet — drop files below to start</span>
      </div>
    </div>
    <div class="session-right">
      <button class="btn" id="import-btn" title="Continue from a previously downloaded Excel">
        <i class="ti ti-file-import"></i> Import Excel
      </button>
      <button class="btn" id="clear-data-btn" style="display:none; color:var(--danger); border-color:rgba(239,68,68,.35);" title="Clear all accumulated rows and start fresh">
        <i class="ti ti-trash-x"></i> Clear Rows
      </button>
      <button class="btn" id="browse-btn">
        <i class="ti ti-folder-open"></i> Browse files
      </button>
    </div>
  </div>

  <!-- Drop zone -->
  <div class="dropzone" id="dz">
    <div class="dz-icon"><i class="ti ti-cloud-upload"></i></div>
    <div class="dz-title">Drop files here to extract</div>
    <div class="dz-sub">Each file is extracted and merged into one growing Excel — download anytime</div>
    <div class="dz-tags">
      <span class="dz-tag">.jpg</span>
      <span class="dz-tag">.jpeg</span>
      <span class="dz-tag">.png</span>
      <span class="dz-tag">.webp</span>
      <span class="dz-tag">.gif</span>
      <span class="dz-tag">.pdf</span>
    </div>
  </div>
  <input type="file" id="file-input"  accept="image/*,.gif,.pdf" multiple style="display:none">
  <input type="file" id="excel-input" accept=".xlsx"        style="display:none">

  <!-- Queue -->
  <div class="queue-wrap" id="queue-wrap" style="display:none">
    <div class="queue-hdr">
      <span class="queue-label">Files</span>
      <button class="clear-btn" id="clear-btn"><i class="ti ti-trash"></i> Clear done</button>
    </div>
    <div class="queue-list" id="queue-list"></div>
  </div>

</main>

<div class="toast" id="toast"><i id="t-icon"></i><span id="t-msg"></span></div>

<script>
  const dz         = document.getElementById('dz');
  const fileInput  = document.getElementById('file-input');
  const excelInput = document.getElementById('excel-input');
  const browseBtn  = document.getElementById('browse-btn');
  const importBtn  = document.getElementById('import-btn');
  const queueWrap  = document.getElementById('queue-wrap');
  const queueList  = document.getElementById('queue-list');
  const clearBtn   = document.getElementById('clear-btn');
  const totalEl    = document.getElementById('total');
  const dlBtn      = document.getElementById('dl-btn');
  const clearDataBtn = document.getElementById('clear-data-btn');
  const sessionSt  = document.getElementById('session-status');
  const sessionLbl = document.getElementById('session-label');
  const toast      = document.getElementById('toast');
  const tIcon      = document.getElementById('t-icon');
  const tMsg       = document.getElementById('t-msg');

  // ── State: the accumulated Excel lives here in the browser ─────────────────
  let excelBlob = null;   // Blob — the client owns the accumulated Excel entirely
  let totalRows = 0;
  let busy      = false;
  const queue    = [];
  const LS_KEY   = 'tex_backup_v1';   // localStorage key for session persistence

  // ── localStorage helpers ───────────────────────────────────────────────────
  async function lsSave(blob, rows) {
    try {
      const ab  = await blob.arrayBuffer();
      const u8  = new Uint8Array(ab);
      let bin   = '';
      // chunk to avoid call-stack overflow on large arrays
      for (let i = 0; i < u8.length; i += 8192)
        bin += String.fromCharCode(...u8.subarray(i, i + 8192));
      localStorage.setItem(LS_KEY, JSON.stringify({ b64: btoa(bin), rows, ts: Date.now() }));
    } catch { /* quota exceeded or private-mode restriction — silently ignore */ }
  }

  function lsRestore() {
    try {
      const raw = localStorage.getItem(LS_KEY);
      if (!raw) return false;
      const { b64, rows, ts } = JSON.parse(raw);
      const bin  = atob(b64);
      const u8   = new Uint8Array(bin.length);
      for (let i = 0; i < bin.length; i++) u8[i] = bin.charCodeAt(i);
      const blob = new Blob([u8], { type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' });
      excelBlob = blob;
      totalRows = rows;
      totalEl.textContent = rows.toLocaleString() + ' rows';
      dlBtn.disabled = false;
      clearDataBtn.style.display = 'inline-flex';
      sessionSt.classList.add('has-data');
      const age = Math.round((Date.now() - ts) / 60000);
      const ageStr = age < 2 ? 'just now' : age < 60 ? `${age}m ago` : `${Math.round(age/60)}h ago`;
      sessionLbl.textContent = `${rows.toLocaleString()} rows restored (saved ${ageStr}) — keep adding files or Download Excel`;
      return true;
    } catch {
      localStorage.removeItem(LS_KEY);
      return false;
    }
  }

  function setExcel(blob, rows) {
    excelBlob = blob;
    totalRows = rows;
    totalEl.textContent = rows.toLocaleString() + ' rows';
    dlBtn.disabled = false;
    clearDataBtn.style.display = 'inline-flex';
    sessionSt.classList.add('has-data');
    sessionLbl.textContent = rows.toLocaleString() + ' rows accumulated — keep adding files or Download Excel';
    lsSave(blob, rows);   // persist to localStorage so tab-close / refresh doesn't lose data
  }

  function clearData() {
    excelBlob = null;
    totalRows = 0;
    totalEl.textContent = '0 rows';
    dlBtn.disabled = true;
    clearDataBtn.style.display = 'none';
    sessionSt.classList.remove('has-data');
    sessionLbl.textContent = 'No data yet — drop files below to start';
    localStorage.removeItem(LS_KEY);
    fetch('/clear', { method: 'POST' }).catch(() => {});  // wipe server-side backup (local only)
  }

  clearDataBtn.addEventListener('click', () => {
    if (!confirm('Clear all accumulated rows and start fresh?')) return;
    clearData();
    showToast('Rows cleared — ready for a fresh extraction', 's');
  });

  // ── Init: restore previous session, detect local vs cloud ──────────────────
  (async () => {
    try {
      const cfg = await fetch('/config').then(r => r.json());
      if (cfg.is_local) document.getElementById('open-btn').style.display = 'inline-flex';
    } catch {}

    // 1. Try localStorage first (works on Railway and local alike)
    lsRestore();

    // 2. On local dev, also try server-side backup — it covers sessions from other
    //    browser tabs or machines and will override the localStorage copy if newer.
    try {
      const res = await fetch('/resume');
      if (res.status === 200) {
        const blob = await res.blob();
        const rows = parseInt(res.headers.get('X-Total-Rows') || '0');
        setExcel(blob, rows);   // also saves the fresher copy back to localStorage
        sessionLbl.textContent = rows.toLocaleString() + ' rows restored from last session';
      }
    } catch {}
  })();

  // ── Download ───────────────────────────────────────────────────────────────
  dlBtn.addEventListener('click', () => {
    if (!excelBlob) return;
    const url = URL.createObjectURL(excelBlob);
    const a   = document.createElement('a');
    a.href     = url;
    a.download = 'extracted.xlsx';
    a.click();
    URL.revokeObjectURL(url);
  });

  // ── Import existing Excel (continue a previous session) ───────────────────
  importBtn.addEventListener('click', () => excelInput.click());
  excelInput.addEventListener('change', e => {
    const f = e.target.files[0];
    if (!f) return;
    excelInput.value = '';
    excelBlob = f;
    sessionSt.classList.add('has-data');
    sessionLbl.textContent = `Imported "${f.name}" — drop more files to append`;
    dlBtn.disabled = false;
    clearDataBtn.style.display = 'inline-flex';
    showToast(`"${f.name}" loaded — ready to append more extractions`, 's');
  });

  // ── Drag & Drop ────────────────────────────────────────────────────────────
  dz.addEventListener('dragover',  e => { e.preventDefault(); dz.classList.add('over'); });
  dz.addEventListener('dragleave', () => dz.classList.remove('over'));
  dz.addEventListener('drop', e => { e.preventDefault(); dz.classList.remove('over'); addFiles([...e.dataTransfer.files]); });
  dz.addEventListener('click', () => fileInput.click());
  browseBtn.addEventListener('click', e => { e.stopPropagation(); fileInput.click(); });
  fileInput.addEventListener('change', e => { addFiles([...e.target.files]); fileInput.value = ''; });

  // ── Queue ──────────────────────────────────────────────────────────────────
  const EXTS = ['.jpg','.jpeg','.png','.webp','.gif','.pdf'];

  function addFiles(files) {
    for (const f of files) {
      const ext = '.' + f.name.split('.').pop().toLowerCase();
      if (!EXTS.includes(ext)) { showToast(`"${f.name}" — unsupported format`, 'e'); continue; }
      const item = { file: f, el: makeItem(f), status: 'pending' };
      queue.push(item);
      queueList.appendChild(item.el);
    }
    queueWrap.style.display = queue.length ? 'flex' : 'none';
    tick();
  }

  function makeItem(f) {
    const pdf = f.name.toLowerCase().endsWith('.pdf');
    const el  = document.createElement('div');
    el.className = 'qi';
    el.innerHTML =
      `<div class="qi-ico ${pdf?'pdf':'img'}"><i class="ti ti-${pdf?'file-type-pdf':'photo'}"></i></div>` +
      `<div class="qi-body"><div class="qi-name">${esc(f.name)}</div><div class="qi-meta">${fmtSize(f.size)}</div></div>` +
      `<div class="qi-stat pending"><i class="ti ti-clock"></i> Pending</div>`;
    return el;
  }

  function setStatus(item, status, label) {
    item.status = status;
    item.el.className = 'qi ' + status;
    const s = item.el.querySelector('.qi-stat');
    s.className = 'qi-stat ' + status;
    // remove any old retry button
    item.el.querySelector('.qi-retry')?.remove();
    if (status === 'processing') {
      s.innerHTML = '<div class="spinner"></div> Extracting...';
    } else if (status === 'done') {
      s.innerHTML = `<i class="ti ti-circle-check"></i> +${label} rows`;
    } else if (status === 'error') {
      s.innerHTML = `<i class="ti ti-alert-circle"></i> Failed`;
      const meta = item.el.querySelector('.qi-meta');
      if (meta) { meta.textContent = label; meta.style.color = 'var(--danger)'; }
      // add retry button
      const btn = document.createElement('button');
      btn.className = 'qi-retry';
      btn.innerHTML = '<i class="ti ti-refresh"></i> Retry';
      btn.style.cssText = 'margin-left:8px;padding:2px 10px;font-size:12px;border:1px solid var(--danger);background:transparent;color:var(--danger);border-radius:4px;cursor:pointer;';
      btn.onclick = () => { setStatus(item, 'pending', ''); tick(); };
      s.after(btn);
    }
  }

  async function tick() {
    if (busy) return;
    const next = queue.find(i => i.status === 'pending');
    if (!next) return;
    busy = true;
    setStatus(next, 'processing');

    try {
      const fd = new FormData();
      fd.append('file', next.file, next.file.name);
      if (excelBlob) fd.append('base', excelBlob, 'base.xlsx');

      const resp = await fetch('/upload', { method: 'POST', body: fd });

      if (resp.ok && resp.headers.get('Content-Type')?.includes('spreadsheetml')) {
        const blob  = await resp.blob();
        const added = parseInt(resp.headers.get('X-Rows-Added') || '0');
        const total = parseInt(resp.headers.get('X-Total-Rows')  || '0');
        setExcel(blob, total);
        setStatus(next, 'done', added);
        showToast(`"${next.file.name}" — ${added} rows appended`, 's');
      } else {
        const data = await resp.json().catch(() => ({ error: 'Server error' }));
        setStatus(next, 'error', data.error || 'Extraction failed');
        showToast(data.error || 'Extraction failed', 'e');
      }
    } catch (e) {
      try { setStatus(next, 'error', e.message); } catch (_) {}
      showToast('Could not reach server', 'e');
    } finally {
      busy = false;
      tick();   // always advance to the next file, no matter what happened
    }
  }

  clearBtn.addEventListener('click', () => {
    queue.filter(i => i.status==='done'||i.status==='error').forEach(i => {
      i.el.remove(); queue.splice(queue.indexOf(i), 1);
    });
    queueWrap.style.display = queue.length ? 'flex' : 'none';
  });

  // ── Toast ──────────────────────────────────────────────────────────────────
  let tt;
  function showToast(msg, type) {
    clearTimeout(tt);
    tMsg.textContent = msg;
    tIcon.className  = type === 's' ? 'ti ti-circle-check t-s' : 'ti ti-alert-triangle t-e';
    toast.className  = `toast ${type} show`;
    tt = setTimeout(() => toast.className = `toast ${type}`, 4500);
  }

  function esc(s)  { return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }
  function fmtSize(b) {
    if (b < 1024)    return b + ' B';
    if (b < 1048576) return (b/1024).toFixed(1) + ' KB';
    return (b/1048576).toFixed(1) + ' MB';
  }
</script>
</body>
</html>"""

# ── Launch ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if IS_LOCAL:
        import threading, webbrowser
        threading.Timer(1.2, lambda: webbrowser.open(f"http://localhost:{PORT}")).start()
        print(f"\n  Table Extractor  ->  http://localhost:{PORT}")
        print(f"  Ctrl+C to stop\n")
    app.run(host="0.0.0.0", port=PORT, debug=False, threaded=True)
