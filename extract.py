"""
Batch column-by-column extractor for Arabic UAE vehicle seizure GIF documents.

Strategy (reduces hallucination):
  • The page has 3 physical table blocks side-by-side (RTL order: right → middle → left).
  • Each block is cropped and sent to Gemini as a separate batch call.
  • Optionally the full page is also sent for cross-checking.
  • Results are merged using serial-number / VIN dedup (column-by-column fill).
  • Output: one clean, sorted, numbered Excel file.
"""

import base64, json, sys, os, time, hashlib, io
from pathlib import Path
from PIL import Image

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Config ────────────────────────────────────────────────────────────────────
TARGET  = Path(r"c:\Users\Abood\Downloads\files (1)\data\3-16444-5_02.gif")
OUTPUT  = Path(r"c:\Users\Abood\Downloads\files (1)\3-16444-5_02_extracted.xlsx")
API_KEY = os.environ.get("GEMINI_API_KEY", "")
MODEL   = "gemini-2.5-flash"

# 8 canonical columns (Arabic RTL document)
COLS = ["م", "اللوحة", "مصدر اللوحة", "فئة اللوحة",
        "نوع المركبة", "لونها", "تاريخ الحجز", "الملاحظات"]

RATE_DELAY  = 5.0   # seconds between Gemini calls (free tier ≤ 15 RPM)
MAX_PX      = 1800  # resize longest side to this before sending
JPEG_Q      = 88
N_DOC_COLS  = 3     # number of physical column-blocks on the page
OVERLAP_PCT = 0.06  # 6% column overlap when cropping

# ── Prompts ───────────────────────────────────────────────────────────────────
BASE_RULES = """
COLUMN CONTRACT — every row MUST have exactly 8 values in THIS order:
  [0] م           → row serial number (integer, e.g. "1", "2", …)
  [1] اللوحة      → plate / registration number (digits only, e.g. "12345")
  [2] مصدر اللوحة → plate source emirate (e.g. "دبي", "أبوظبي", "الشارقة")
  [3] فئة اللوحة  → plate category code (short, e.g. "1", "2", "A", "ب")
  [4] نوع المركبة → vehicle make/model (Arabic text, e.g. "تويوتا كامري")
  [5] لونها       → vehicle colour (Arabic, e.g. "أبيض", "أسود", "فضي")
  [6] تاريخ الحجز → seizure date (e.g. "15/01/2024" or "2024/01/15")
  [7] الملاحظات   → VIN / chassis number (10-22 alphanumeric chars with Latin letters)

RULES:
  • Return ONLY a JSON array of arrays — zero markdown, zero extra keys.
  • Each element is an 8-string array: ["م_value", "plate", "source", …, "VIN"]
  • Use "" for blank or unreadable cells. NEVER skip a row.
  • Do NOT translate — keep original Arabic / numbers exactly.
  • Arabic tables read RIGHT-TO-LEFT: the rightmost column is م.
  • Rows are ordered by the م (serial) column ascending.
"""

def _full_prompt() -> str:
    return (
        "You are a precise Arabic table extractor.\n\n"
        "This image shows a complete page with a table split into "
        f"{N_DOC_COLS} physical column-blocks side-by-side (RTL order: right, middle, left).\n"
        "Reconstruct ONE unified table by reading all blocks and ordering rows by م (serial number).\n\n"
        + BASE_RULES
    )

# ── Image helpers ─────────────────────────────────────────────────────────────
def gif_frames(path: Path) -> list:
    img = Image.open(path)
    frames, seen = [], set()
    try:
        while True:
            f = img.convert("RGB")
            thumb = f.copy(); thumb.thumbnail((16, 16))
            h = hashlib.md5(thumb.tobytes()).hexdigest()
            if h not in seen:
                seen.add(h); frames.append(f.copy())
            img.seek(img.tell() + 1)
    except EOFError:
        pass
    img.close()
    return frames


def to_b64(img: Image.Image) -> tuple:
    w, h = img.size
    if max(w, h) > MAX_PX:
        scale = MAX_PX / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=JPEG_Q)
    return base64.b64encode(buf.getvalue()).decode(), "image/jpeg"


def horizontal_halves(frame: Image.Image) -> list:
    """Split frame into top and bottom halves with 8% overlap.

    Horizontal splits keep ALL columns visible in each batch — unlike vertical
    strips which cut columns apart and produce unresolvable partial rows.
    """
    W, H = frame.size
    mid = H // 2
    ovlp = int(H * 0.04)   # 4% vertical overlap
    top = frame.crop((0, 0, W, min(H, mid + ovlp)))
    bot = frame.crop((0, max(0, mid - ovlp), W, H))
    return [top, bot]

# ── Gemini API ────────────────────────────────────────────────────────────────
import urllib.request, urllib.error

def call_gemini(b64: str, mime: str, prompt: str, label: str) -> str:
    url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
           f"{MODEL}:generateContent?key={API_KEY}")
    payload = json.dumps({
        "contents": [{"parts": [
            {"text": prompt},
            {"inlineData": {"mimeType": mime, "data": b64}},
        ]}],
        "generationConfig": {
            "responseMimeType": "application/json",
            "temperature": 0,
            "thinkingConfig": {"thinkingBudget": 2048},
        },
    }).encode()
    req = urllib.request.Request(
        url, data=payload,
        headers={"Content-Type": "application/json"}, method="POST")
    for attempt in range(1, 5):
        try:
            with urllib.request.urlopen(req, timeout=240) as resp:
                return resp.read().decode()
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            if e.code in (429, 500, 503) and attempt < 4:
                wait = 40 * attempt
                print(f"    [{label}] HTTP {e.code} — retrying in {wait}s...")
                time.sleep(wait)
            else:
                raise RuntimeError(f"[{label}] HTTP {e.code}: {body[:300]}")
    raise RuntimeError(f"[{label}] All retries failed")

# ── Parse Gemini response → list of 8-element rows ───────────────────────────
def parse_rows(raw: str, label: str) -> list:
    try:
        data = json.loads(raw)
        text = data["candidates"][0]["content"]["parts"][0]["text"].strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
        parsed = json.loads(text)
    except Exception as e:
        print(f"    [{label}] parse error: {e}")
        return []

    rows = []
    if isinstance(parsed, list):
        for item in parsed:
            if isinstance(item, list):
                rows.append(item)
            elif isinstance(item, dict):
                rows.extend(item.get("rows", []))
    return rows

# ── Column-by-column dedup merge ─────────────────────────────────────────────
SI, PI, VI = 0, 1, 7   # serial, plate, VIN column indices

def _key(row: list) -> str:
    def g(i): return str(row[i] if i < len(row) else "").strip()
    # Plate is the most reliable cross-strip key: each vehicle has a unique plate,
    # and different strips may give the same row with different amounts of data filled in.
    # Using plate first ensures both "partial" and "full" extractions of the same row merge.
    plate = g(PI)
    if plate.isdigit() and 2 <= len(plate) <= 9:
        return f"p:{plate}"
    # VIN as fallback when plate is missing/invalid
    vin = g(VI)
    if len(vin) >= 10 and any(c.isdigit() for c in vin) and any(c.isascii() and c.isalpha() for c in vin):
        return f"v:{vin}"
    # Serial last resort (risky: different strips restart serial from 1 for their block)
    serial = g(SI)
    if serial.isdigit():
        return f"s:{serial}"
    return ""


def col_merge(raw_rows: list) -> list:
    """
    Merge all raw rows column-by-column:
    • Identify each row by VIN → serial → plate.
    • For each unique key, keep the first non-empty value per column.
    • Rows with no reliable key are kept once but cannot be deduped.
    """
    store: dict = {}
    order: list = []
    no_key = 0

    for row in raw_rows:
        row = (list(row) + [""] * 8)[:8]   # normalise to exactly 8 cells
        key = _key(row)
        if not key:
            no_key += 1
            key = f"_:{no_key}"
        if key not in store:
            store[key] = [""] * 8
            order.append(key)
        bucket = store[key]
        for i in range(8):
            v = str(row[i] or "").strip()
            if v and not bucket[i]:
                bucket[i] = v

    result = [store[k] for k in order if any(store[k])]
    result.sort(key=lambda r: (0, int(r[0])) if r[0].isdigit() else (1, 0))
    return result

# ── Excel writer ──────────────────────────────────────────────────────────────
def write_excel(rows: list, path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.rightToLeft = True

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin      = Side(style="thin", color="CCCCCC")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Header row
    for ci, h in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = ctr;  cell.border = bdr
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # Data rows
    data_font = Font(name="Calibri", size=10)
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")   # light blue alternate rows
    for ri, row in enumerate(rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.font = data_font; cell.border = bdr; cell.alignment = ctr
            if fill:
                cell.fill = fill

    # Column widths
    for ci in range(1, len(COLS) + 1):
        vals = [ws.cell(row=r, column=ci).value or "" for r in range(1, ws.max_row + 1)]
        w = max((len(str(v)) for v in vals), default=8)
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 42)

    wb.save(str(path))

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print(f"  File  : {TARGET.name}")
    print(f"  Output: {OUTPUT.name}")
    print("=" * 60)

    if not TARGET.exists():
        sys.exit(f"ERROR: file not found — {TARGET}")

    frames = gif_frames(TARGET)
    print(f"\n  Unique GIF frames: {len(frames)}")
    for fi, frm in enumerate(frames):
        print(f"    Frame {fi+1}: {frm.size[0]}×{frm.size[1]} px")

    all_raw: list = []
    call_count = 0

    for fi, frame in enumerate(frames, 1):
        W, H = frame.size
        print(f"\n── Frame {fi}/{len(frames)} ({W}×{H}) ──")

        # Send the full frame — keeps all columns together so Gemini can read
        # every row completely without guessing missing columns.
        # A second identical call is made as a cross-check: if both attempts
        # agree on row count the result is reliable; the dedup merges them.
        for attempt in range(1, 3):
            label = f"frame{fi}-pass{attempt}"
            b64, mime = to_b64(frame)
            kb = len(base64.b64decode(b64)) // 1024
            print(f"  [{label}] {W}×{H} → {kb} KB (pass {attempt}/2)")

            if call_count > 0:
                print(f"  Waiting {RATE_DELAY}s...")
                time.sleep(RATE_DELAY)

            try:
                raw  = call_gemini(b64, mime, _full_prompt(), label)
                rows = parse_rows(raw, label)
                print(f"  [{label}] → {len(rows)} rows")
                all_raw.extend(rows)
                call_count += 1
            except Exception as e:
                print(f"  [{label}] ERROR: {e}")

    print(f"\n── Merge ──")
    print(f"  Total raw rows (with duplicates): {len(all_raw)}")
    merged = col_merge(all_raw)
    print(f"  After column-by-column dedup    : {len(merged)} unique rows")

    if not merged:
        print("\nWARNING: No rows extracted. Check API key and image quality.")
        return

    # Renumber م column sequentially to fix any gaps / duplicates
    for i, row in enumerate(merged, 1):
        row[SI] = str(i)

    print(f"\n── Writing Excel ──")
    write_excel(merged, OUTPUT)
    print(f"  Done: {len(merged)} rows → {OUTPUT}")
    print("=" * 60)


if __name__ == "__main__":
    main()
